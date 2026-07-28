import os
import streamlit as st
from openpyxl import load_workbook
from db import get_conn, ph


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(BASE_DIR, "Master Lists.xlsx")


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_active(value):
    text = _normalize_text(value).upper()
    if text in ("Y", "YES", "TRUE", "1"):
        return "Y"
    return "N"


def _normalize_flag(value):
    text = _normalize_text(value).upper()
    if text in ("Y", "YES", "TRUE", "1"):
        return "Y"
    return "N"


def _normalize_role(value):
    text = _normalize_text(value).upper()
    allowed = {"TEAM", "STORE", "LINREP", "LINTEAM", "LINSUP", "ADMIN", "BOSS"}
    if text not in allowed:
        raise ValueError(f"Invalid role in User Master: {text}")
    return text


def _open_workbook():
    if not os.path.exists(MASTER_FILE):
        raise FileNotFoundError(f"Workbook not found: {MASTER_FILE}")
    return load_workbook(MASTER_FILE, data_only=True)


# ---------------------------
# User Master
# ---------------------------
def load_user_master_rows():
    wb = _open_workbook()
    if "User Master" not in wb.sheetnames:
        raise ValueError("Sheet 'User Master' not found in Master Lists.xlsx")

    ws = wb["User Master"]
    headers = [_normalize_text(cell.value) for cell in ws[1]]

    required = ["Name", "Username", "Password", "Role", "Group", "Active"]
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing required column in User Master: {col}")

    idx = {name: headers.index(name) for name in headers}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _normalize_text(row[idx["Name"]])
        username = _normalize_text(row[idx["Username"]])
        password = _normalize_text(row[idx["Password"]])
        role = _normalize_role(row[idx["Role"]])
        team_code = _normalize_text(row[idx["Group"]])
        active = _normalize_active(row[idx["Active"]])

        if not any([name, username, password, role, team_code]):
            continue

        if not username:
            raise ValueError("A row in User Master is missing Username.")
        if not name:
            raise ValueError(f"User Master row for username '{username}' is missing Name.")
        if not password:
            raise ValueError(f"User Master row for username '{username}' is missing Password.")
        if not team_code:
            raise ValueError(f"User Master row for username '{username}' is missing Group.")

        rows.append({
            "display_name": name,
            "username": username,
            "raw_password": password,
            "role": role,
            "team_code": team_code,
            "active": active,
        })

    return rows


def sync_users_from_workbook():
    from auth import hash_password  # local import avoids circular import

    user_rows = load_user_master_rows()

    conn = get_conn()
    cur = conn.cursor()

    for u in user_rows:
        query = f"SELECT user_id FROM users WHERE username = {ph()}"
        cur.execute(query, (u["username"],))
        existing = cur.fetchone()

        password_hash = hash_password(u["raw_password"])

        if existing:
            cur.execute(f"""
                UPDATE users
                SET display_name = {ph()},
                    password_hash = {ph()},
                    role = {ph()},
                    team_code = {ph()},
                    active = {ph()}
                    WHERE username = {ph()}
            """, (
                u["display_name"],
                password_hash,
                u["role"],
                u["team_code"],
                u["active"],
                u["username"],
            ))
        else:
            cur.execute(f"""
                INSERT INTO users (
                    username,
                    display_name,
                    password_hash,
                    role,
                    team_code,
                    active
                )
                VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()})
            """, (
                u["username"],
                u["display_name"],
                password_hash,
                u["role"],
                u["team_code"],
                u["active"],
            ))

    conn.commit()
    conn.close()


# ---------------------------
# Item Master
# ---------------------------
@st.cache_data(show_spinner=False)
def load_item_master_rows():
    wb = _open_workbook()
    if "Item Master" not in wb.sheetnames:
        raise ValueError("Sheet 'Item Master' not found in Master Lists.xlsx")

    ws = wb["Item Master"]
    headers = [_normalize_text(cell.value) for cell in ws[1]]

    required = [
        "Item Name",
        "Unit",
        "Note",
        "display_order",
        "Tue",
        "Fri",
        "OT",
        "Adhoc",
        "Inventory",
    ]
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing required column in Item Master: {col}")

    idx = {name: headers.index(name) for name in headers}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_name = _normalize_text(row[idx["Item Name"]])
        unit = _normalize_text(row[idx["Unit"]])
        note = _normalize_text(row[idx["Note"]])
        display_order_raw = row[idx["display_order"]]
        tue = _normalize_flag(row[idx["Tue"]])
        fri = _normalize_flag(row[idx["Fri"]])
        ot = _normalize_flag(row[idx["OT"]])
        adhoc = _normalize_flag(row[idx["Adhoc"]])
        inventory = _normalize_flag(row[idx["Inventory"]])
        category = _normalize_text(row[idx["Category"]]) if "Category" in idx else "Others"
        if not category:
            category = "Others"

        if not any([item_name, unit, note, display_order_raw, tue, fri, ot, adhoc, inventory]):
            continue

        if not item_name:
            raise ValueError("A row in Item Master is missing Item Name.")
        if display_order_raw in (None, ""):
            raise ValueError(f"Item Master row '{item_name}' is missing display_order.")

        try:
            display_order = int(display_order_raw)
        except Exception:
            raise ValueError(f"Item Master row '{item_name}' has invalid display_order: {display_order_raw}")

        rows.append({
            "item_name": item_name,
            "unit": unit,
            "note": note,
            "display_order": display_order,
            "Tue": tue,
            "Fri": fri,
            "OT": ot,
            "Adhoc": adhoc,
            "Inventory": inventory,
            "category": category,
        })

    rows.sort(key=lambda x: x["display_order"])
    return rows


def get_item_master_lookup():
    rows = load_item_master_rows()
    lookup = {}

    for r in rows:
        lookup[r["item_name"]] = {
            "unit": r["unit"],
            "note": r["note"],
            "category": r.get("category", "Others"),
            "display_order": r["display_order"],
            "Tue": r["Tue"],
            "Fri": r["Fri"],
            "OT": r["OT"],
            "Adhoc": r["Adhoc"],
            "Inventory": r["Inventory"],
        }

    return lookup


# ---------------------------
# Report Line Master
# ---------------------------
@st.cache_data(show_spinner=False)
def load_report_line_master_rows():
    """
    Reads the 'Report Line Master' sheet from Master Lists.xlsx.

    Expected columns:
    - report_line_id
    - for_column_b
    - report_line_name
    - report_uom
    - for_column_e
    - for_column_f
    - display_line_order
    """
    wb = _open_workbook()
    if "Report Line Master" not in wb.sheetnames:
        raise ValueError("Sheet 'Report Line Master' not found in Master Lists.xlsx")

    ws = wb["Report Line Master"]
    headers = [_normalize_text(cell.value) for cell in ws[1]]

    required = [
        "report_line_id",
        "for_column_b",
        "report_line_name",
        "report_uom",
        "for_column_e",
        "for_column_f",
        "display_line_order",
    ]
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing required column in Report Line Master: {col}")

    idx = {name: headers.index(name) for name in headers}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        report_line_id = _normalize_text(row[idx["report_line_id"]])
        for_column_b = _normalize_text(row[idx["for_column_b"]])
        report_line_name = _normalize_text(row[idx["report_line_name"]])
        report_uom = _normalize_text(row[idx["report_uom"]])
        for_column_e = _normalize_text(row[idx["for_column_e"]])
        for_column_f = _normalize_text(row[idx["for_column_f"]])
        display_line_order_raw = row[idx["display_line_order"]]

        if not any([
            report_line_id,
            for_column_b,
            report_line_name,
            report_uom,
            for_column_e,
            for_column_f,
            display_line_order_raw,
        ]):
            continue

        if not report_line_id:
            raise ValueError("A row in Report Line Master is missing report_line_id.")
        if not report_line_name:
            raise ValueError(f"Report Line Master row '{report_line_id}' is missing report_line_name.")
        if display_line_order_raw in (None, ""):
            raise ValueError(f"Report Line Master row '{report_line_id}' is missing display_line_order.")

        try:
            display_line_order = int(display_line_order_raw)
        except Exception:
            raise ValueError(
                f"Report Line Master row '{report_line_id}' has invalid display_line_order: {display_line_order_raw}"
            )

        rows.append({
            "report_line_id": report_line_id,
            "for_column_b": for_column_b,
            "report_line_name": report_line_name,
            "report_uom": report_uom,
            "for_column_e": for_column_e,
            "for_column_f": for_column_f,
            "display_line_order": display_line_order,
        })

    rows.sort(key=lambda x: x["display_line_order"])
    return rows


# ---------------------------
# Report Mapping
# ---------------------------
@st.cache_data(show_spinner=False)
def load_report_mapping_rows():
    """
    Reads the 'Report Mapping' sheet.

    Expected columns:
    - item_name
    - report_line_name

    Returns rows in the form:
    [
        {
            "item_name": ...,
            "report_line_name": ...
        },
        ...
    ]
    """
    wb = _open_workbook()
    if "Report Mapping" not in wb.sheetnames:
        raise ValueError("Sheet 'Report Mapping' not found in Master Lists.xlsx")

    ws = wb["Report Mapping"]
    headers = [_normalize_text(cell.value) for cell in ws[1]]

    required = ["item_name", "report_line_name"]
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing required column in Report Mapping: {col}")

    idx = {name: headers.index(name) for name in headers}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_name = _normalize_text(row[idx["item_name"]])
        report_line_name = _normalize_text(row[idx["report_line_name"]])

        # skip fully blank rows
        if not any([item_name, report_line_name]):
            continue

        # skip partially filled rows instead of crashing
        if not item_name or not report_line_name:
            continue

        rows.append({
            "item_name": item_name,
            "report_line_name": report_line_name,
        })

    return rows

@st.cache_data(show_spinner=False)
def load_destinations_rows():
    """
    Reads the 'Destinations' sheet from Master Lists.xlsx.

    Expected columns:
    - destination_name
    - display_order
    - active
    - report_group

    Returns rows sorted by display_order.
    """
    wb = _open_workbook()
    if "Destinations" not in wb.sheetnames:
        raise ValueError("Sheet 'Destinations' not found in Master Lists.xlsx")

    ws = wb["Destinations"]
    headers = [_normalize_text(cell.value) for cell in ws[1]]

    required = [
        "destination_name",
        "display_order",
        "active",
        "report_group",
    ]
    for col in required:
        if col not in headers:
            raise ValueError(f"Missing required column in Destinations: {col}")

    idx = {name: headers.index(name) for name in headers}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        destination_name = _normalize_text(row[idx["destination_name"]])
        display_order_raw = row[idx["display_order"]]
        active = _normalize_active(row[idx["active"]])
        report_group = _normalize_text(row[idx["report_group"]])

        if not any([destination_name, display_order_raw, active, report_group]):
            continue

        if not destination_name:
            continue

        if display_order_raw in (None, ""):
            raise ValueError(f"Destination '{destination_name}' is missing display_order.")

        try:
            display_order = int(display_order_raw)
        except Exception:
            raise ValueError(
                f"Destination '{destination_name}' has invalid display_order: {display_order_raw}"
            )

        rows.append({
            "destination_name": destination_name,
            "display_order": display_order,
            "active": active,
            "report_group": report_group,
        })

    rows.sort(key=lambda x: x["display_order"])
    return rows

@st.cache_data(show_spinner=False)
def load_audit_locations_rows():
    """
    Reads 'Glo Gel Locations' sheet.

    Returns:
    - structured list of locations
    """
    wb = _open_workbook()
    ws = wb["Glo Gel Locations"]

    headers = [_normalize_text(cell.value) for cell in ws[1]]
    idx = {name: headers.index(name) for name in headers}

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        location_name = _normalize_text(row[idx["location_name"]])
        if not location_name:
            continue

        rows.append({
            "location_name": location_name,
            "tower": _normalize_text(row[idx["tower"]]),
            "level": row[idx["level"]],
            "zone": row[idx.get("zone")] if "zone" in idx else None,
            "display_order": int(row[idx["display_order"]]),
            "active": _normalize_active(row[idx["active"]]),

            # group visibility
            "for_AB2-B1": _normalize_active(row[idx["for_AB2-B1"]]),
            "for_AA1-3": _normalize_active(row[idx["for_AA1-3"]]),
            "for_A4-8": _normalize_active(row[idx["for_A4-8"]]),
            "for_B1-4": _normalize_active(row[idx["for_B1-4"]]),
            "for_B5-10": _normalize_active(row[idx["for_B5-10"]]),
            "for_B11-16": _normalize_active(row[idx["for_B11-16"]]),
            "for_C1-12": _normalize_active(row[idx["for_C1-12"]]),
        })

    rows.sort(key=lambda x: x["display_order"])
    return rows

@st.cache_data(show_spinner=False)
def load_audit_surfaces_rows():
    """
    Reads 'Audit Surfaces' sheet.

    Returns:
    - surface definitions for both templates
    """
    wb = _open_workbook()
    ws = wb["Audit Surfaces"]

    headers = [_normalize_text(cell.value) for cell in ws[1]]
    idx = {name: headers.index(name) for name in headers}

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        surface_id = _normalize_text(row[idx["surface_id"]])
        surface_name = _normalize_text(row[idx["surface_name"]])

        if not surface_id or not surface_name:
            continue

        rows.append({
            "surface_id": surface_id,
            "surface_name": surface_name,
            "active": _normalize_active(row[idx["active"]]),

            # AC (Tower A / C)
            "in_ac": _normalize_active(row[idx["in_ac"]]),
            "ac_area_group": _normalize_text(row[idx["ac_area_group"]]),
            "ac_display_order": row[idx["ac_display_order"]],

            # B (Tower B)
            "in_b": _normalize_active(row[idx["in_b"]]),
            "b_area_group": _normalize_text(row[idx["b_area_group"]]),
            "b_display_order": row[idx["b_display_order"]],
        })

    return rows

@st.cache_data
def load_linen_locations_rows():
    wb = _open_workbook()
    ws = wb["Linen Location Master"]

    headers = [_normalize_text(cell.value) for cell in ws[1]]
    idx = {name: headers.index(name) for name in headers}

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        location_id = _normalize_text(row[idx["location_id"]])
        location_name = _normalize_text(row[idx["location_name"]])

        tower = _normalize_text(row[idx["tower"]])
        level = _normalize_text(row[idx["level"]])
        zone = _normalize_text(row[idx["zone"]])

        lin_linrep = _normalize_text(row[idx["lin_LINREP"]]) if "lin_LINREP" in idx else ""

        lin_b1_4 = _normalize_text(row[idx["lin_B1-4"]]) if "lin_B1-4" in idx else ""
        lin_b5_10 = _normalize_text(row[idx["lin_B5-10"]]) if "lin_B5-10" in idx else ""
        lin_b11_16 = _normalize_text(row[idx["lin_B11-16"]]) if "lin_B11-16" in idx else ""
        lin_c1_12 = _normalize_text(row[idx["lin_C1-12"]]) if "lin_C1-12" in idx else ""

        if not location_id:
            continue

        rows.append({
            "location_id": location_id,
            "tower": tower,
            "level": level,
            "zone": zone,
            "location_name": location_name,
            "lin_LINREP": lin_linrep,
            "lin_B1-4": lin_b1_4,
            "lin_B5-10": lin_b5_10,
            "lin_B11-16": lin_b11_16,
            "lin_C1-12": lin_c1_12,
        })

    return rows

def get_linen_location_map():
    locations = load_linen_locations_rows()

    return {
        row["location_id"]: row
        for row in locations
    }

@st.cache_data
def load_linen_master_rows():
    wb = _open_workbook()
    ws = wb["Linen Master"]

    headers = [_normalize_text(cell.value) for cell in ws[1]]
    idx = {name: headers.index(name) for name in headers}

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item_no = _normalize_text(row[idx["item_no"]])
        item_name = _normalize_text(row[idx["item_name"]])
        lin_category = _normalize_text(row[idx["lin_category"]])

        if not item_no or not item_name:
            continue

        item = {
            "item_no": item_no,
            "item_name": item_name,
            "lin_category": lin_category,
        }

        for h in headers:
            if h.startswith("LOC"):
                item[h] = _normalize_flag(row[idx[h]])

        rows.append(item)

    return rows

def get_linen_items_for_location(location_id):
    rows = load_linen_master_rows()

    return [
        r for r in rows
        if r.get(location_id) == "Y"
    ]