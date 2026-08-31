from db import get_conn, ph
from report_config import get_report_lines, get_report_line_id_for_item
from master_loader import load_destinations_rows, MASTER_FILE
from openpyxl import load_workbook
from fractions import Fraction


# ---------------------------
# Period helpers
# ---------------------------
def get_half_year_months(period_code: str):
    """
    H1 = Jan to Jun
    H2 = Jul to Dec
    """
    period_code = period_code.upper()

    if period_code == "H1":
        return [1, 2, 3, 4, 5, 6]
    elif period_code == "H2":
        return [7, 8, 9, 10, 11, 12]
    else:
        raise ValueError("period_code must be 'H1' or 'H2'")


# ---------------------------
# Destination / bucket helpers
# ---------------------------
def _get_destination_group_lookup():
    """
    Returns:
        {
            "Annex": "ANX Blk",
            "Laundry Bay": "Others",
            ...
        }
    """
    rows = load_destinations_rows()
    return {
        r["destination_name"].strip(): r["report_group"]
        for r in rows
        if r["active"] == "Y"
    }


def get_order_bucket(team_code: str, template_day: str):
    """
    Determines report sheet bucket for normal order-issued movements.
    """
    if template_day in ("ANNEX_TUE", "ANNEX_FRI"):
        return "ANX Blk"

    # follow workbook/live naming as source of truth
    if team_code in ("AB2-B1", "AA1-3", "A4-8"):
        return "Tower A"

    if team_code in ("B1-4", "B5-10", "B11-16"):
        return "Tower B"

    if team_code in ("C1-12",):
        return "Tower C"

    return None


# ---------------------------
# Main report builder
# ---------------------------
def get_half_year_report_data(year: int, period_code: str):
    """
    Builds half-year monthly report data.

    Includes:
    - ORDER-issued stock movements
    - ADHOC issue stock movements

    Returns:
    {
        "Tower ABC": [...],
        "Tower A": [...],
        "Tower B": [...],
        "Tower C": [...],
        "ANX Blk": [...],
        "Others": [...],
    }
    """
    months = get_half_year_months(period_code)
    destination_group_lookup = _get_destination_group_lookup()

    conn = get_conn()
    cur = conn.cursor()

    # Pull all OUT movements for the selected year
    query = f"""
        SELECT
            sm.item_name,
            sm.qty,
            sm.created_at,
            sm.source_type,
            sm.source_id,
            sm.issued_to,
            o.team_code,
            o.template_day
        FROM stock_movements sm
        LEFT JOIN orders o
            ON sm.source_type = 'ORDER'
           AND sm.source_id = o.order_id
        WHERE sm.movement_type = 'OUT'
          AND COALESCE(sm.is_voided, FALSE) = FALSE
          AND EXTRACT(YEAR FROM sm.created_at::timestamp) = {ph()}
    """
    cur.execute(query, (year,))

    rows = cur.fetchall()
    conn.close()

    # Filter to selected half-year months
    filtered_rows = []
    for row in rows:
        created_at = row["created_at"]  # YYYY-MM-DD HH:MM:SS
        month = int(created_at[5:7])
        if month in months:
            filtered_rows.append(dict(row))

    report_lines = get_report_lines()

    bucket_names = ["Tower A", "Tower B", "Tower C", "ANX Blk", "Others", "Tower ABC"]

    # bucket -> report_line_id -> month -> qty
    bucket_monthly_totals = {}
    for bucket in bucket_names:
        bucket_monthly_totals[bucket] = {}
        for line in report_lines:
            rid = line["report_line_id"]
            bucket_monthly_totals[bucket][rid] = {m: 0 for m in months}

    # Fill tower-specific buckets
    for row in filtered_rows:
        item_name = row["item_name"]
        qty = int(row["qty"] or 0)
        month = int(row["created_at"][5:7])
        source_type = row["source_type"]

        report_line_id = get_report_line_id_for_item(item_name)
        if not report_line_id:
            continue

        bucket = None

        if source_type == "ORDER":
            team_code = row["team_code"]
            template_day = row["template_day"]
            bucket = get_order_bucket(team_code, template_day)

        elif source_type in ("ADHOC", "STOCK_ISSUE"):
            issued_to = (row["issued_to"] or "").strip()
            bucket = destination_group_lookup.get(issued_to, "Others")

        if not bucket:
            continue

        if bucket not in bucket_monthly_totals:
            continue

        bucket_monthly_totals[bucket][report_line_id][month] += qty

    # Build Tower ABC as sum of A + B + C + ANX Blk
    # (Others is intentionally NOT included in Tower ABC)
    for line in report_lines:
        rid = line["report_line_id"]
        for month in months:
            bucket_monthly_totals["Tower ABC"][rid][month] = (
                bucket_monthly_totals["Tower A"][rid][month]
                + bucket_monthly_totals["Tower B"][rid][month]
                + bucket_monthly_totals["Tower C"][rid][month]
                + bucket_monthly_totals["ANX Blk"][rid][month]
                + bucket_monthly_totals["Others"][rid][month]
            )

    # Convert into ordered row structures with static A:F fields
    result = {}
    for bucket in bucket_names:
        result[bucket] = []
        for line in report_lines:
            rid = line["report_line_id"]
            result[bucket].append({
                "report_line_id": line["report_line_id"],     # A
                "for_column_b": line["for_column_b"],         # B
                "report_line_name": line["report_line_name"], # C
                "report_uom": line["report_uom"],             # D
                "for_column_e": line["for_column_e"],         # E
                "for_column_f": line["for_column_f"],         # F
                "monthly_qty": bucket_monthly_totals[bucket][rid],  # G:L
            })

    return result



# ---------------------------
# App K2 report builder
# ---------------------------
def _parse_k2_conversion_factor(value):
    """
    Accepts normal numeric factors (1, 0.1, 12) and text fractions
    such as '1/12'. Returns an exact Fraction where possible.
    """
    if value in (None, ""):
        return Fraction(1, 1)

    if isinstance(value, int):
        return Fraction(value, 1)

    if isinstance(value, float):
        return Fraction(str(value))

    text = str(value).strip()
    try:
        return Fraction(text)
    except Exception as exc:
        raise ValueError(f"Invalid App K2 conversion_factor: {value}") from exc


def _load_app_k2_mapping_rows():
    """Reads active rows from the 'App K2 Mapping' master-list sheet."""
    wb = load_workbook(MASTER_FILE, data_only=True, read_only=True)
    if "App K2 Mapping" not in wb.sheetnames:
        wb.close()
        raise ValueError("Sheet 'App K2 Mapping' not found in Master Lists.xlsx")

    ws = wb["App K2 Mapping"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    required = [
        "app_K2_report_line_id",
        "app_K2_report_line_name",
        "item_name_for_app_K2",
        "app_K2_for_column_b",
        "app_K2_for_column_d",
        "app_K2_for_column_e",
        "app_K2_for_column_f",
        "app_K2_for_column_G",
        "conversion_factor",
        "active",
    ]
    for col in required:
        if col not in headers:
            wb.close()
            raise ValueError(f"Missing required column in App K2 Mapping: {col}")

    idx = {name: headers.index(name) for name in headers}
    rows = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        active = str(values[idx["active"]] or "").strip().upper()
        if active != "Y":
            continue

        line_id_raw = values[idx["app_K2_report_line_id"]]
        if line_id_raw in (None, ""):
            continue

        try:
            line_id = int(line_id_raw)
        except Exception as exc:
            wb.close()
            raise ValueError(f"Invalid app_K2_report_line_id: {line_id_raw}") from exc

        rows.append({
            "line_id": line_id,
            "line_name": str(values[idx["app_K2_report_line_name"]] or "").strip(),
            "item_name": str(values[idx["item_name_for_app_K2"]] or "").strip(),
            "column_b": str(values[idx["app_K2_for_column_b"]] or "").strip(),
            "column_d": str(values[idx["app_K2_for_column_d"]] or "").strip(),
            "column_e": str(values[idx["app_K2_for_column_e"]] or "").strip(),
            "column_f": str(values[idx["app_K2_for_column_f"]] or "").strip(),
            "column_g": values[idx["app_K2_for_column_G"]],
            "factor": _parse_k2_conversion_factor(values[idx["conversion_factor"]]),
        })

    wb.close()
    rows.sort(key=lambda x: x["line_id"])
    return rows


def get_app_k2_report_data(year: int, period_code: str):
    """
    Builds the 'For App K2' worksheet data from non-voided Stock In movements.

    Multiple mapping rows may share the same App K2 line ID. Each source item's
    Stock In quantity is converted first, then all contributions are summed.
    """
    months = get_half_year_months(period_code)
    mapping_rows = _load_app_k2_mapping_rows()

    conn = get_conn()
    cur = conn.cursor()
    query = f"""
        SELECT item_name, qty, created_at
        FROM stock_movements
        WHERE movement_type = 'IN'
          AND COALESCE(is_voided, FALSE) = FALSE
          AND EXTRACT(YEAR FROM created_at::timestamp) = {ph()}
    """
    cur.execute(query, (year,))
    movement_rows = cur.fetchall()
    conn.close()

    # item -> month -> exact integer stock-in quantity
    item_month_totals = {}
    for row in movement_rows:
        created_at = row["created_at"]
        if hasattr(created_at, "month"):
            month = int(created_at.month)
        else:
            month = int(str(created_at)[5:7])

        if month not in months:
            continue

        item_name = str(row["item_name"] or "").strip()
        qty = int(row["qty"] or 0)
        item_month_totals.setdefault(item_name, {m: 0 for m in months})
        item_month_totals[item_name][month] += qty

    # Preserve one output row per App K2 line ID while allowing many source items.
    line_map = {}
    for mapping in mapping_rows:
        line_id = mapping["line_id"]
        if line_id not in line_map:
            line_map[line_id] = {
                "line_id": line_id,
                "line_name": mapping["line_name"],
                "column_b": mapping["column_b"],
                "column_d": mapping["column_d"],
                "column_e": mapping["column_e"],
                "column_f": mapping["column_f"],
                "column_g": mapping["column_g"],
                "monthly_qty": {m: Fraction(0, 1) for m in months},
            }

        item_name = mapping["item_name"]
        if not item_name:
            continue

        source_totals = item_month_totals.get(item_name, {})
        factor = mapping["factor"]
        for month in months:
            line_map[line_id]["monthly_qty"][month] += Fraction(
                int(source_totals.get(month, 0)), 1
            ) * factor

    result = []
    for line_id in sorted(line_map):
        row = line_map[line_id]
        display_monthly = {}
        for month, qty in row["monthly_qty"].items():
            display_monthly[month] = int(qty) if qty.denominator == 1 else float(qty)
        row["monthly_qty"] = display_monthly
        result.append(row)

    return result
