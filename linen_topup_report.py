import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from linen_topup_db import build_manual_topup_month_report


# Excel worksheet tab colours
TAB_COLOURS = {
    "A": "4472C4",   # Blue
    "B": "70AD47",   # Green
    "C": "ED7D31",   # Orange
}


def generate_manual_topup_report(year, month):
    year = int(year)
    month = int(month)

    days_in_month = calendar.monthrange(
        year,
        month
    )[1]

    report = build_manual_topup_month_report(
        year,
        month
    )

    wb = Workbook()

    # Remove default Sheet.
    default_ws = wb.active
    wb.remove(default_ws)

    # ---------------------------
    # Common formatting
    # ---------------------------
    thin = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    header_font = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    left = Alignment(
        horizontal="left",
        vertical="center"
    )

    month_name = calendar.month_name[month]

    # ---------------------------
    # Create worksheets
    # ---------------------------
    for sheet_name, sheet_data in report.items():

        # Excel worksheet names cannot exceed 31 characters.
        excel_sheet_name = str(sheet_name)[:31]

        ws = wb.create_sheet(
            title=excel_sheet_name
        )

        tower = str(
            sheet_data.get("tower") or ""
        ).strip().upper()

        if tower in TAB_COLOURS:
            ws.sheet_properties.tabColor = (
                TAB_COLOURS[tower]
            )

        # ---------------------------
        # Title
        # ---------------------------
        total_columns = (
            2
            + days_in_month
            + 1
        )

        last_column = get_column_letter(
            total_columns
        )

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=total_columns,
        )

        title_cell = ws.cell(
            row=1,
            column=1
        )

        title_cell.value = (
            f"{sheet_name} - "
            f"Manual Top Up "
            f"{month_name} {year}"
        )

        title_cell.font = Font(
            bold=True,
            size=14
        )

        title_cell.alignment = center

        # ---------------------------
        # Headers
        # ---------------------------
        header_row = 3

        ws.cell(
            header_row,
            1,
            "S/N"
        )

        ws.cell(
            header_row,
            2,
            "ITEM"
        )

        for day in range(
            1,
            days_in_month + 1
        ):
            ws.cell(
                header_row,
                day + 2,
                day
            )

        total_col = days_in_month + 3

        ws.cell(
            header_row,
            total_col,
            "Total Top Up"
        )

        for col in range(
            1,
            total_col + 1
        ):
            cell = ws.cell(
                header_row,
                col
            )

            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center

        # ---------------------------
        # Report lines 1-39
        # ---------------------------
        first_data_row = header_row + 1

        for report_no in range(1, 40):

            row_no = (
                first_data_row
                + report_no
                - 1
            )

            row_data = sheet_data[
                "rows"
            ][report_no]

            ws.cell(
                row_no,
                1,
                report_no
            )

            ws.cell(
                row_no,
                2,
                row_data["name"]
            )

            ws.cell(
                row_no,
                1
            ).alignment = center

            ws.cell(
                row_no,
                2
            ).alignment = left

            total_quantity = 0

            for day in range(
                1,
                days_in_month + 1
            ):

                quantity = int(
                    row_data[
                        "days"
                    ].get(day, 0)
                    or 0
                )

                total_quantity += quantity

                cell = ws.cell(
                    row_no,
                    day + 2
                )

                # Requirement:
                # zeroes appear blank.
                if quantity > 0:
                    cell.value = quantity
                else:
                    cell.value = None

                cell.alignment = center

            total_cell = ws.cell(
                row_no,
                total_col
            )

            if total_quantity > 0:
                total_cell.value = total_quantity
            else:
                total_cell.value = None

            total_cell.alignment = center

            # Borders for whole row
            for col in range(
                1,
                total_col + 1
            ):
                ws.cell(
                    row_no,
                    col
                ).border = border

        # ---------------------------
        # Widths
        # ---------------------------
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 34

        for day_col in range(
            3,
            total_col
        ):
            letter = get_column_letter(
                day_col
            )
            ws.column_dimensions[
                letter
            ].width = 5

        ws.column_dimensions[
            get_column_letter(total_col)
        ].width = 14

        # ---------------------------
        # Freeze panes / filters
        # ---------------------------
        ws.freeze_panes = "C4"

        ws.auto_filter.ref = (
            f"A{header_row}:"
            f"{last_column}"
            f"{first_data_row + 38}"
        )

        # Print setup
        ws.sheet_view.showGridLines = False

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.print_title_rows = (
            f"1:{header_row}"
        )

        ws.print_area = (
            f"A1:"
            f"{last_column}"
            f"{first_data_row + 38}"
        )

    # ---------------------------
    # Save to memory
    # ---------------------------
    output = BytesIO()

    wb.save(output)
    output.seek(0)

    return output