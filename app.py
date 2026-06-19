import streamlit as st
from datetime import date
from orders_db import cancel_order

from audit_db import (
    get_visible_locations_for_user,
    get_visible_zones_for_user,
    get_visible_locations_for_user_and_zone,
    get_template_for_tower,
    get_surfaces_for_template,
    list_audits_for_user,
    create_audit_header,
    get_audit_header,
    get_audit_results,
    save_audit_detail,
    get_completed_audits_grouped_by_tower,
    get_surface_template_for_tower,
)

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from master_loader import (
    get_item_master_lookup,
    load_destinations_rows,
    load_linen_locations_rows,
    get_linen_location_map,
    get_linen_items_for_location,
    load_linen_master_rows,
)

from admin_tools import reset_orders_only, reset_orders_and_movements

from packing_db import get_packing_list_data, save_packing_board_issued

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime

from report_db import get_half_year_report_data

from report_db import get_half_year_report_data

from stock_db import (
    create_stock_in,
    get_inventory_rows,
    create_adhoc_issue_batch,
    search_stock_movements,
    void_stock_movement,
    get_stock_card_rows,
    save_opening_balance_override,
    update_stock_movement_date
)

from io import BytesIO
from openpyxl import Workbook
from db import init_db, seed_minimal_data
from auth import authenticate
from run_dates import compute_run_date
from orders_db import (
    list_orders_for_store,
    list_orders_for_team,
    get_order_by_unique,
    get_order,
    create_order,
    get_order_lines,
    update_order_lines,
)

from linen_db import (
    create_linen_cycle,
    get_linen_cycles,
    get_linen_cycle,
    get_cycle_reps,
    save_cycle_reps,
    get_cycle_assignments,
    save_cycle_assignments,
    start_linen_cycle,
    get_active_linen_cycle,
    get_assignments_for_user,
    get_submission,
    get_submission_lines,
    save_submission_draft,
    submit_submission,
    get_submission_status_map,
    complete_linen_cycle,
    get_submitted_location_count,
    get_cycle_submission_lines,
    force_complete_linen_cycle,
    get_linen_rep_names
)


st.set_page_config(page_title="Ops Hub", layout="centered")


# ---------------------------
# App bootstrap
# ---------------------------
def ensure_bootstrap():
    init_db()
    seed_minimal_data()


# ---------------------------
# Session helpers
# ---------------------------
def logout():
    st.session_state.pop("user", None)
    st.session_state["page"] = "login"


def require_login():
    if "user" not in st.session_state or not st.session_state["user"]:
        st.session_state["page"] = "login"
        st.stop()


# ---------------------------
# UI helpers
# ---------------------------
def format_status(status: str) -> str:
    if status == "PENDING":
        return "🟡 PENDING"
    if status == "PARTIALLY_ISSUED":
        return "🟠 PARTIALLY ISSUED"
    if status == "ISSUED":
        return "🟢 ISSUED"
    return status

def can_cancel_order(user_role, username, order_creator, status):

    if status != "Pending":
        return False

    if user_role in ["ADMIN", "STORE"]:
        return True

    if username == order_creator:
        return True

    return False

# ---------------------------
# Login page
# ---------------------------
def page_login():
    st.title("Ops Hub")
    st.markdown(
        "<p style='font-size:14px; color:gray;'>"
        "Housekeeping & Linen Operations System"
        "</p>",
        unsafe_allow_html=True
    )
    st.markdown(
        "<p style='font-size:18px; font-weight:600;'>Login</p>",
        unsafe_allow_html=True
    )

    username = st.text_input("Username", placeholder="e.g. HenryC")
    pin = st.text_input("PIN / Password", type="password", placeholder="Enter password")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Log in", use_container_width=True):
            user = authenticate(username.strip(), pin.strip())
            if user:
                st.session_state["user"] = user
                st.session_state["page"] = "home"
                st.rerun()
            else:
                st.error("Invalid username or password.")
    with col2:
        if st.button("Clear", use_container_width=True):
            st.rerun()
    
    st.divider()
    st.markdown("""
    ### Announcements

    19/6/2026
    - Linen Inventory is now live

    16/6/2026
    - Revised Stock Card reporting and Glo Gel Audit filters
    - Added Stock-In date selector (STORE)
    - New Packing List format (STORE)

    2/6/2026
    - Fixed report grouping and Stock Issue logic mapping
    - Fixed Stock Card Export excel syntax issues
    
    """)


# ---------------------------
# Home page (role-based)
# ---------------------------
def page_home():
    require_login()
    user = st.session_state["user"]
    role = user["role"]

    active_linen_cycle = get_active_linen_cycle()
    linen_active = active_linen_cycle is not None

    st.title("Home")
    st.write(f"Logged in as **{user['username']}** ({role})")

    # TEAM
    if role == "TEAM":
        if st.button("Orders", use_container_width=True):
            st.session_state["page"] = "orders"
            st.rerun()

        if st.button("Glo Gel Audit", use_container_width=True):
            st.session_state["page"] = "glo_gel_audits"
            st.rerun()

        if user["team_code"] in ("B1-4", "B5-10", "B11-16", "C1-12"):
            if st.button(
                "Linen Inventory",
                use_container_width=True,
                disabled=not linen_active
            ):
                st.session_state["page"] = "linen_inventory"
                st.rerun()

            if not linen_active:
                st.caption("No active linen inventory session.")

        st.button("Logout", use_container_width=True, on_click=logout)
        return

    # STORE / ADMIN
    if role in ("STORE", "ADMIN"):
        if st.button("Orders", use_container_width=True):
            st.session_state["page"] = "orders"
            st.rerun()

        if st.button("Issue Stock", use_container_width=True):
            st.session_state["page"] = "issue_stock"
            st.rerun()

        if st.button("Stock In", use_container_width=True):
            st.session_state["page"] = "stock_in"
            st.rerun()

        if st.button("Inventory", use_container_width=True):
            st.session_state["page"] = "inventory"
            st.rerun()

        if st.button("Monthly Report", use_container_width=True):
            st.session_state["page"] = "monthly_report"
            st.rerun()

        if st.button("Stock Card Export", use_container_width=True):
            st.session_state["page"] = "stock_card"
            st.rerun()

        if role == "ADMIN":
            if st.button("System Tools", use_container_width=True):
                st.session_state["page"] = "system_tools"
                st.rerun()

            if st.button("Glo Gel Audit", use_container_width=True):
                st.session_state["page"] = "glo_gel_audits"
                st.rerun()

        if role == "ADMIN":

            if st.button("Linen Inventory", use_container_width=True):
                st.session_state["page"] = "linen_inventory"
                st.rerun()

        st.button("Logout", use_container_width=True, on_click=logout)
        return

    # LINEN ROLES
    if role == "LINSUP":
        if st.button("Initiate/ Manage Linen Inventory", use_container_width=True):
            st.session_state["page"] = "linen_manage"
            st.rerun()
    
    if role in ("LINSUP", "LINTEAM", "LINREP"):
        if st.button(
            "Linen Inventory",
            use_container_width=True,
            disabled=not linen_active
        ):
            st.session_state["page"] = "linen_inventory"
            st.rerun()

        if not linen_active:
            st.caption("No active linen inventory session.")

        st.button("Logout", use_container_width=True, on_click=logout)
        return

    if role == "LINREP":
        active_cycle = get_active_linen_cycle()

        if active_cycle:
            rep_names = get_linen_rep_names(active_cycle["id"])

            assigned_name = rep_names.get(
                user["username"],
                "Not Assigned"
            )

            st.caption(
                f"Assigned Linen Rep: {assigned_name}"
            )

    # BOSS
    if role == "BOSS":
        if st.button("Monthly Report", use_container_width=True):
            st.session_state["page"] = "monthly_report"
            st.rerun()

        if st.button("Stock Card Export", use_container_width=True):
            st.session_state["page"] = "stock_card"
            st.rerun()

        st.button("Logout", use_container_width=True, on_click=logout)
        return

# ---------------------------
# Orders page router
# ---------------------------
def page_orders():
    require_login()
    user = st.session_state["user"]
    role = user["role"]

    st.title("Orders")

    if role == "STORE":
        page_orders_store(user)
        return

    if role == "TEAM":
        page_orders_team(user)
        return

    if role == "ADMIN":
        page_orders_admin(user)
        return

    st.error("Access denied.")


# ---------------------------
# Orders page: STORE
# ---------------------------
def page_orders_store(user):
    st.caption("Storeman view: pending orders first.")

    status_filter = st.selectbox(
        "Filter",
        ["PENDING", "PARTIALLY_ISSUED", "ISSUED", "ALL"],
        index=0,
    )
    status = None if status_filter == "ALL" else status_filter

    orders = list_orders_for_store(status=status)

    if not orders:
        st.info("No orders found.")
    else:
        for o in orders:
            header = f"{format_status(o['status'])} | {o['team_code']} | {o['template_day']} | Issue {o['run_date']} | by {o['created_by']}"

            with st.expander(header, expanded=False):
                st.write(f"Submitted: **{o['created_at']}**")
                if o.get("issued_at"):
                    st.write(f"Last issued: **{o['issued_at']}** by **{o.get('issued_by', '-') }**")

                if st.button(
                    "Open Order",
                    key=f"store_open_{o['order_id']}",
                    use_container_width=True,
                ):
                    
                    st.session_state["active_order_id"] = o["order_id"]
                    st.session_state["page"] = "order_detail"
                    st.rerun()

                if o["status"] == "PENDING":

                    reason = st.text_input(
                        "Cancellation Reason",
                        key=f"store_cancel_reason_{o['order_id']}"
                    )

                    if st.button(
                        "Cancel Order",
                        key=f"store_cancel_{o['order_id']}",
                        use_container_width=True
                    ):

                        if reason.strip():

                            rows_updated = cancel_order(
                                o["order_id"],
                                user["username"],
                                reason.strip()
                            )

                            if rows_updated > 0:
                                st.success("Order cancelled.")
                                st.rerun()
                            else:
                                st.warning("No order was cancelled. It may already be issued/cancelled, or the status did not match.")

                        else:
                            st.warning("Please enter a cancellation reason.")

    st.divider()
    st.subheader("Packing List Export")

    c1, c2 = st.columns(2)

    with c1:
        if st.button("Generate Tuesday Packing List", use_container_width=True):
            try:
                st.session_state["packing_output"] = generate_packing_list_excel(mode="TUE_COMBINED")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    with c2:
        if st.button("Generate Friday Packing List", use_container_width=True):
            try:
                st.session_state["packing_output"] = generate_packing_list_excel(mode="FRI_COMBINED")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    if "packing_output" in st.session_state:
        st.download_button(
            label="Download Packing List",
            data=st.session_state["packing_output"],
            file_name="packing_list.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Back", use_container_width=True):
            st.session_state["page"] = "home"
            st.rerun()
    with c2:
        st.button("Logout", use_container_width=True, on_click=logout)


# ---------------------------
# Orders page: TEAM
# ---------------------------
def page_orders_team(user):
    st.caption("Team view: one order per issue date. Only creator can edit pending.")

    orders = list_orders_for_team(user["team_code"])

    if not orders:
        st.info("No orders yet.")
    else:
        for o in orders:
            header = f"{format_status(o['status'])} | {o['template_day']} | Issue {o['run_date']} | by {o['created_by']}"

            with st.expander(header, expanded=False):
                st.write(f"Group: **{o['team_code']}**")
                st.write(f"Created: **{o['created_at']}**")

                if o.get("issued_at"):
                    st.write(f"Last issued: **{o['issued_at']}** by **{o.get('issued_by', '-') }**")

                is_creator = (o["created_by"] == user["username"])
                can_edit = (o["status"] == "PENDING") and is_creator

                c1, c2 = st.columns([1, 2])
                with c1:
                    if st.button(
                        "Edit Order" if can_edit else "View Order",
                        key=f"team_open_{o['order_id']}",
                        use_container_width=True,
                    ):
                        st.session_state["active_order_id"] = o["order_id"]
                        st.session_state["page"] = "order_detail"
                        st.rerun()
                with c2:
                    if not is_creator and o["status"] == "PENDING":
                        st.caption("Edit locked: only the creator can edit this pending order.")
                    elif o["status"] in ("PARTIALLY_ISSUED", "ISSUED"):
                        st.caption("This order is no longer editable by team users.")

                if can_edit:

                    reason = st.text_input(
                        "Cancellation Reason",
                        key=f"team_cancel_reason_{o['order_id']}"
                    )

                    if st.button(
                        "Cancel Order",
                        key=f"team_cancel_{o['order_id']}",
                        use_container_width=True
                    ):

                        if reason.strip():

                            rows_updated = cancel_order(
                                o["order_id"],
                                user["username"],
                                reason.strip()
                            )

                            if rows_updated > 0:
                                st.success("Order cancelled.")
                                st.rerun()
                            else:
                                st.warning("No order was cancelled. It may already be issued/cancelled, or the status did not match.")

                        else:
                            st.warning("Please enter a cancellation reason.")

    st.divider()
    st.subheader("Create / Open Order")

    # Row 1
    c1, c2 = st.columns(2)

    with c1:
        if st.button("Tuesday Order", use_container_width=True):
            _open_or_create_order("TUE")

    with c2:
        if st.button("Friday Order", use_container_width=True):
            _open_or_create_order("FRI")

    # Row 2
    c3, c4 = st.columns(2)

    with c3:
        if user["team_code"] == "AA1-3":
            if st.button("Annex - Tuesday", use_container_width=True):
                _open_or_create_order("ANNEX_TUE")

    with c4:
        if user["team_code"] == "AA1-3":
            if st.button("Annex - Friday", use_container_width=True):
                _open_or_create_order("ANNEX_FRI")

    # OT stays separate (optional row 3)
    if user["team_code"] == "B1-4":
        st.divider()
        if st.button("OT Order", use_container_width=True):
            _open_or_create_order("OT")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Back", use_container_width=True):
            st.session_state["page"] = "home"
            st.rerun()
    with c2:
        st.button("Logout", use_container_width=True, on_click=logout)


# ---------------------------
# Orders page: ADMIN
# ---------------------------
def page_orders_admin(user):
    st.caption("Admin view: browse all orders.")

    orders = list_orders_for_store(status=None)

    if not orders:
        st.info("No orders found.")
    else:
        for o in orders:
            header = f"{format_status(o['status'])} | {o['team_code']} | {o['template_day']} | Issue {o['run_date']}"

            with st.expander(header, expanded=False):
                if st.button(
                    "Open Order",
                    key=f"admin_open_{o['order_id']}",
                    use_container_width=True,
                ):
                    st.session_state["active_order_id"] = o["order_id"]
                    st.session_state["page"] = "order_detail"
                    st.rerun()

            if o["status"] == "PENDING":

                reason = st.text_input(
                    "Cancellation Reason",
                    key=f"admin_cancel_reason_{o['order_id']}"
                )

                if st.button(
                    "Cancel Order",
                    key=f"admin_cancel_{o['order_id']}",
                    use_container_width=True
                ):

                    if reason.strip():

                        rows_updated = cancel_order(
                            o["order_id"],
                            user["username"],
                            reason.strip()
                        )

                        if rows_updated > 0:
                            st.success("Order cancelled.")
                            st.rerun()
                        else:
                            st.warning("No order was cancelled. It may already be issued/cancelled, or the status did not match.")

                    else:
                        st.warning("Please enter a cancellation reason.")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Back", use_container_width=True):
            st.session_state["page"] = "home"
            st.rerun()
    with c2:
        st.button("Logout", use_container_width=True, on_click=logout)


# ---------------------------
# Create / open order helper
# ---------------------------
def _open_or_create_order(template_day: str):
    user = st.session_state["user"]
    team_code = user["team_code"]

    today = date.today()

    if template_day == "OT":
        # OT is special and can be made any day; for now use today's date as issue date
        run_date = today.isoformat()
    else:
        run_date = compute_run_date(today, template_day).isoformat()

    existing = get_order_by_unique(team_code, template_day, run_date)

    if existing:
        st.session_state["active_order_id"] = existing["order_id"]
        st.session_state["page"] = "order_detail"
        st.rerun()
    else:
        new_id = create_order(
            team_code=team_code,
            template_day=template_day,
            run_date=run_date,
            created_by=user["username"],
        )
        st.session_state["active_order_id"] = new_id
        st.session_state["page"] = "order_detail"
        st.rerun()


# ---------------------------
# Order detail page
# ---------------------------
def page_order_detail():
    require_login()
    user = st.session_state["user"]

    if user["role"] == "BOSS":
        st.error("Access denied.")
        return

    order_id = st.session_state.get("active_order_id")
    if not order_id:
        st.session_state["page"] = "orders"
        st.rerun()

    o = get_order(int(order_id))
    if not o:
        st.error("Order not found.")
        st.session_state["page"] = "orders"
        st.rerun()

    lines = get_order_lines(int(order_id), template_day=o["template_day"])

    if user["role"] in ("STORE", "ADMIN"):
        lines = [r for r in lines if int(r["qty_requested"] or 0) > 0]

    st.title("Order Detail")
    st.caption(
        f"{o['team_code']} | {o['template_day']} | Issue {o['run_date']} | "
        f"Status: {format_status(o['status'])} | Created by {o['created_by']}"
    )

    is_team_creator = (user["role"] == "TEAM") and (o["created_by"] == user["username"])
    can_edit_request = (user["role"] == "TEAM") and is_team_creator and (o["status"] == "PENDING")

    # STORE / ADMIN can increase issued quantities
    can_edit_issued = user["role"] in ("STORE", "ADMIN") and o["status"] in ("PENDING", "PARTIALLY_ISSUED")

    item_lookup = get_item_master_lookup()

    data = []
    for r in lines:
        item_name = r["item_name"]
        unit = ""
        details = ""
        category = "Others"

        if item_name in item_lookup:
            unit = item_lookup[item_name]["unit"]
            details = item_lookup[item_name]["note"]
            category = item_lookup[item_name].get("category", "Others") or "Others"

        data.append({
            "line_id": r["line_id"],
            "Category": category,
            "Item": item_name,
            "UOM": unit,
            "Details": details,
            "Request": int(r["qty_requested"] or 0),
            "Issued": int(r["qty_issued"] or 0),
        })

    data_key = f"order_detail_data_{order_id}"

    data.sort(key=lambda x: (x.get("Category", "Others"), x.get("Item", "")))

    if data_key not in st.session_state:
        st.session_state[data_key] = data

    original_by_line_id = {int(r["line_id"]): r for r in lines}

    st.subheader("Items")

    disabled_cols = ["No", "UOM", "Details"]
    if not can_edit_request:
        disabled_cols.append("Request")
    if not can_edit_issued:
        disabled_cols.append("Issued")

    if user["role"] in ("STORE", "ADMIN"):
        if st.button("Fulfill All", use_container_width=True, key=f"fulfill_all_{order_id}"):
            fulfilled_rows = []
            for row in st.session_state[data_key]:
                row = dict(row)
                row["Issued"] = int(row["Request"] or 0)
                fulfilled_rows.append(row)

            st.session_state[data_key] = fulfilled_rows
            st.rerun()

    # Group rows by category
    grouped = {}
    for row in st.session_state[data_key]:
        category = row.get("Category", "Others") or "Others"
        grouped.setdefault(category, []).append(row)

    edited_all = []

    for category, rows in grouped.items():
        st.markdown(f"### {category}")

        edited_part = st.data_editor(
            rows,
            hide_index=True,
            disabled=disabled_cols + ["line_id", "Category"],
            use_container_width=True,
            column_config={
                "line_id": None,
                "Category": None,
                "Item": st.column_config.TextColumn("Item", disabled=True),
                "UOM": st.column_config.TextColumn("UOM", disabled=True),
                "Details": st.column_config.TextColumn("Details", disabled=True),
                "Request": st.column_config.NumberColumn("Request", min_value=0, step=1),
                "Issued": st.column_config.NumberColumn("Issued", min_value=0, step=1),
            },
            key=f"editor_{order_id}_{category}",
        )

        edited_all.extend(edited_part)

    edited = edited_all

    show_save = can_edit_request or can_edit_issued

    if show_save:
        save_label = "Save Issued" if user["role"] in ("STORE", "ADMIN") else "Save Request"

        if st.button(save_label, key=f"save_{order_id}_{user['role']}", use_container_width=True):
            rows_to_save = []

            for row in edited:
                lid = int(row.get("line_id") or 0)
                orig = original_by_line_id.get(lid)

                if not orig:
                    continue

                req = int(row["Request"])
                iss = int(row["Issued"])

                if not can_edit_request:
                    req = int(orig["qty_requested"] or 0)

                if not can_edit_issued:
                    iss = int(orig["qty_issued"] or 0)

                # Team users cannot issue. Store/Admin can only increase issued quantity.
                prev_issued = int(orig["qty_issued"] or 0)
                if can_edit_issued and iss < prev_issued:
                    iss = prev_issued

                if iss > req:
                    iss = req

                rows_to_save.append({
                    "line_id": lid,
                    "qty_requested": req,
                    "qty_issued": iss,
                })

            try:
                update_order_lines(
                    int(order_id),
                    rows_to_save,
                    updated_by=user["username"]
                )
                st.success(
                    "Issued quantities saved."
                    if user["role"] in ("STORE", "ADMIN")
                    else "Request saved."
                )

                if data_key in st.session_state:
                    del st.session_state[data_key]

                st.rerun()
            except ValueError as e:
                st.error(str(e))
    else:
        st.caption("No editable fields for your role on this order.")

    st.divider()

    if o["status"] == "ISSUED":
        if user["role"] in ("STORE", "ADMIN"):
            st.info("This order is fully issued. Further increases may still be saved if needed.")
        else:
            st.info("This order is issued and locked.")
    elif o["status"] == "PARTIALLY_ISSUED":
        if user["role"] in ("STORE", "ADMIN"):
            st.info("This order is partially issued. Continue issuing as needed.")
        else:
            st.info("This order is partially issued and locked for team users.")
    else:
        if can_edit_request:
            st.success("You can edit Request quantities.")
        elif user["role"] == "TEAM":
            st.warning("Read-only: only the creator can edit this pending order.")
        if can_edit_issued:
            st.info("Store/Admin can edit Issued quantities.")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Back to Orders", use_container_width=True):
            st.session_state["page"] = "orders"
            st.rerun()
    with c2:
        st.button("Logout", use_container_width=True, on_click=logout)


def page_issue_stock():
    require_login()
    user = st.session_state["user"]

    if user["role"] not in ("STORE", "ADMIN"):
        st.error("Access denied.")
        return

    st.title("Issue Stock")

    if "flash_message" in st.session_state:
        st.success(st.session_state["flash_message"])
        del st.session_state["flash_message"]

    destination_rows = load_destinations_rows()
    active_destinations = [
        r["destination_name"]
        for r in destination_rows
        if r["active"] == "Y"
    ]

    issued_to = st.selectbox("Issue To", active_destinations)
    final_destination = issued_to

    rows = get_inventory_rows()

    if not rows:
        st.info("No items found.")
        return

    data = []
    for r in rows:
        data.append({
            "Item": r["item_name"],
            "Current": int(r["current_stock"]),
            "Issue Qty": "",
        })

    edited = st.data_editor(
        data,
        hide_index=True,
        use_container_width=True,
        column_config={
            "Item": st.column_config.TextColumn("Item", disabled=True),
            "Current": st.column_config.NumberColumn("Current", disabled=True),
            "Issue Qty": st.column_config.TextColumn("Issue Qty"),
        },
        key="issue_stock_editor",
    )

    if st.button("Issue", use_container_width=True):
        if not final_destination:
            st.warning("Please select or specify a destination.")
            return

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "home"
        st.rerun()

    issue_rows = []
    count = 0

    for row in edited:
        raw_qty = str(row["Issue Qty"]).strip()

        if raw_qty == "":
            continue

        try:
            qty = int(raw_qty.replace(",", ""))
        except Exception:
            st.error(f"Invalid Issue Qty for item: {row['Item']}")
            return

        if qty < 0:
            st.error(f"Negative Issue Qty is not allowed for item: {row['Item']}")
            return

        if qty > 0:
            issue_rows.append({
                "item_name": row["Item"],
                "qty": qty,
            })
            count += 1

    if count == 0:
        st.warning("No quantities entered.")
        return

    try:
        create_adhoc_issue_batch(
            issue_rows=issue_rows,
            issued_to=final_destination,
            created_by=user["username"],
        )
        st.session_state["flash_message"] = f"Stock issued for {count} item(s) to {final_destination}."
        st.rerun()
    except ValueError as e:
        st.error(str(e))

    st.divider()


def page_stock_in():
    require_login()
    user = st.session_state["user"]

    if user["role"] not in ("STORE", "ADMIN"):
        st.error("Access denied.")
        return

    st.title("Stock In")

    stock_in_date = st.date_input(
        "Stock In Date",
        value=date.today()
    )

    if "flash_message" in st.session_state:
        st.success(st.session_state["flash_message"])
        del st.session_state["flash_message"]

    rows = get_inventory_rows()

    if not rows:
        st.info("No items found.")
        return

    data = []
    for r in rows:
        data.append({
            "Category": r.get("category") or r.get("Category") or "Others",
            "Item": r["item_name"],
            "Current": int(r["current_stock"]),
            "Stock In": ""
        })

    grouped = {}
    for row in data:
        category = row.get("Category", "Others") or "Others"
        grouped.setdefault(category, []).append(row)

    edited_all = []

    for category, rows in grouped.items():
        st.markdown(f"### {category}")

        edited_part = st.data_editor(
            rows,
            hide_index=True,
            use_container_width=True,
            column_config={
                "Category": None,
                "Item": st.column_config.TextColumn("Item", disabled=True),
                "Current": st.column_config.NumberColumn("Current", disabled=True),
                "Stock In": st.column_config.TextColumn("Stock In"),
            },
            key=f"stock_in_editor_{category}",
        )

        edited_all.extend(edited_part)

    edited = edited_all

    stock_in_date = st.date_input(
        "Stock In Date",
        value=date.today(),
        key="stock_in_date"
    )

    if st.button("Stock In", use_container_width=True):
        count = 0

        for row in edited:
            raw_qty = str(row["Stock In"]).strip()

            if raw_qty == "":
                continue

            try:
                qty = int(raw_qty.replace(",", ""))
            except Exception:
                st.error(f"Invalid Stock In quantity for item: {row['Item']}")
                return

            if qty < 0:
                st.error(f"Negative Stock In quantity is not allowed for item: {row['Item']}")
                return

            if qty == 0:
                continue

            create_stock_in(
                item_name=row["Item"],
                qty=qty,
                created_by=user["username"],
                created_at=str(stock_in_date)
            )
            count += 1

        if count == 0:
            st.warning("No quantities entered.")
        else:
            st.session_state["flash_message"] = f"Stock-in recorded for {count} item(s)."
            st.rerun()

    st.divider()

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "home"
        st.rerun()

def page_inventory():
    require_login()
    user = st.session_state["user"]

    if user["role"] not in ("STORE", "ADMIN"):
        st.error("Access denied.")
        return

    st.title("Inventory")

    rows = get_inventory_rows()

    if not rows:
        st.info("No inventory data.")
        return

    data = []
    for r in rows:
        data.append({
            "Category": r.get("category", "Others"),
            "Item": r["item_name"],
            "Current Stock": int(r["current_stock"])
        })

    grouped = {}

    for row in data:
        category = row.get("Category", "Others") or "Others"
        grouped.setdefault(category, []).append(row)

    for category, rows in grouped.items():
        st.markdown(f"### {category}")

        st.dataframe(
            rows,
            hide_index=True,
            use_container_width=True,
            column_config={
                "Category": None,
            },
        )

    st.divider()

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "home"
        st.rerun()


def page_monthly_report():
    require_login()
    user = st.session_state["user"]

    if user["role"] not in ("STORE", "ADMIN", "BOSS"):
        st.error("Access denied.")
        return

    st.title("Monthly Report")
    st.caption("Export fixed half-year stock records template.")

    year = date.today().year

    c1, c2 = st.columns(2)

    with c1:
        if st.button("H1 (Jan - Jun)", use_container_width=True):
            _generate_half_year_report_excel(year, "H1")

    with c2:
        if st.button("H2 (Jul - Dec)", use_container_width=True):
            _generate_half_year_report_excel(year, "H2")

    st.divider()

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "home"
        st.rerun()

def blank_if_zero(value):
    if value in (0, 0.0, "0", "0.0"):
        return ""
    return value

def _generate_half_year_report_excel(year: int, period_code: str):
    report_data = get_half_year_report_data(year, period_code)

    wb = Workbook()

    # remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    if period_code == "H1":
        months = [1, 2, 3, 4, 5, 6]
    else:
        months = [7, 8, 9, 10, 11, 12]

    month_headers = [date(year, m, 1).strftime("%b %y") for m in months]

    sheet_order = ["Tower ABC", "Tower A", "Tower B", "Tower C", "ANX Blk", "Others"]

    for sheet_name in sheet_order:
        ws = wb.create_sheet(title=sheet_name)

        # Top title
        ws["A1"] = "Stock Records"
        ws["A2"] = sheet_name
        ws["A3"] = f"{period_code} {year}"

        # Static headers A:F
        ws["A5"] = "SN"
        ws["B5"] = "Category"
        ws["C5"] = "Description"
        ws["D5"] = "UOM"
        ws["E5"] = ""
        ws["F5"] = "Unit Rate"

        # Dynamic month headers G:L
        start_col = 7  # G
        for i, header in enumerate(month_headers):
            ws.cell(row=5, column=start_col + i, value=header)

        excel_row = 6

        for row in report_data[sheet_name]:
            # Static columns A:F
            ws.cell(row=excel_row, column=1, value=row["report_line_id"])     # A
            ws.cell(row=excel_row, column=2, value=row["for_column_b"])       # B
            ws.cell(row=excel_row, column=3, value=row["report_line_name"])   # C
            ws.cell(row=excel_row, column=4, value=row["report_uom"])         # D
            ws.cell(row=excel_row, column=5, value=row["for_column_e"])       # E
            ws.cell(row=excel_row, column=6, value=row["for_column_f"])       # F

            # Month columns G:L
            monthly_qty = row["monthly_qty"]
            for i, month in enumerate(months):
                ws.cell(row=excel_row, column=start_col + i, value=blank_if_zero(monthly_qty.get(month, 0)))

            excel_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"stock_records_{period_code}_{year}.xlsx"

    st.download_button(
        label=f"Download {period_code} Report Excel",
        data=output,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

def safe_sheet_name(name):
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']

    safe = str(name) if name else "Item"

    for ch in invalid_chars:
        safe = safe.replace(ch, "-")

    safe = safe.strip()

    if not safe:
        safe = "Item"

    return safe[:31]

def page_stock_card():
    require_login()
    user = st.session_state["user"]

    if user["role"] not in ("STORE", "ADMIN", "BOSS"):
        st.error("Access denied.")
        return

    st.title("Stock Card Export")

    rows = get_inventory_rows()

    if not rows:
        st.info("No inventory items found.")
        return

    st.caption("Select items to include in the stock card export.")

    selected_month = st.date_input(
        "Stock Card Month",
        value=date.today().replace(day=1)
    )

    date_from = selected_month.replace(day=1)

    if selected_month.month == 12:
        date_to = selected_month.replace(
            year=selected_month.year + 1,
            month=1,
            day=1
        )
    else:
        date_to = selected_month.replace(
            month=selected_month.month + 1,
            day=1
        )

    selections = []
    for idx, r in enumerate(rows):
        col1, col2, col3 = st.columns([1, 6, 2])
        with col1:
            checked = st.checkbox("", key=f"stock_card_item_{idx}")
        with col2:
            st.write(r["item_name"])
        with col3:
            st.write(r["current_stock"])
        if checked:
            selections.append(r["item_name"])

    if st.button("Generate Stock Card Excel", use_container_width=True):
        if not selections:
            st.warning("Please select at least one item.")
            return

        wb = Workbook()
        # remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        for item_name in selections:
            sheet_name = safe_sheet_name(item_name)
            ws = wb.create_sheet(title=sheet_name)

            ws["A1"] = "Item"
            ws["B1"] = item_name

            ws["A3"] = "Date"
            ws["B3"] = "Stock In"
            ws["C3"] = "Stock Out"
            ws["D3"] = "Balance"
            ws["E3"] = "Issued To"
            ws["F3"] = "Remarks"

            stock_rows = get_stock_card_rows(
                item_name=item_name,
                date_from=str(date_from),
                date_to=str(date_to),
            )

            excel_row = 4
            for row in stock_rows:
                ws.cell(row=excel_row, column=1, value=row["Date"])
                ws.cell(row=excel_row, column=2, value=row["Stock In"])
                ws.cell(row=excel_row, column=3, value=row["Stock Out"])
                ws.cell(row=excel_row, column=4, value=row["Balance"])
                ws.cell(row=excel_row, column=5, value=row["Issued To"])
                ws.cell(row=excel_row, column=6, value=row["Remarks"])
                excel_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="Download Stock Card Excel",
            data=output,
            file_name="stock_card_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "home"
        st.rerun()

def generate_packing_list_excel(mode="TUE_COMBINED"):
    data = get_packing_list_data(mode=mode)

    items = data["items"]
    teams = data["team_order"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"

    # -----------------------
    # PAGE SETUP (A3 Landscape)
    # -----------------------
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    # -----------------------
    # HEADER
    # -----------------------
    ws["A1"] = "Packing List"
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    ws["A1"].font = Font(size=14, bold=True)

    # -----------------------
    # TABLE HEADER
    # -----------------------
    headers = ["Item", "Stock"] + teams + ["Remarks"]

    # --------------------------
    # TITLE
    # --------------------------

    title_map = {
        "TUE_COMBINED": "Packing List - Tuesday",
        "FRI_COMBINED": "Packing List - Friday",
        "TUE": "Packing List - Tuesday",
        "FRI": "Packing List - Friday",
        "ANNEX_TUE": "Packing List - Annex Tuesday",
        "ANNEX_FRI": "Packing List - Annex Friday",
    }

    title = title_map.get(mode, "Packing List")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(teams))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(size=16, bold=True)
    cell.alignment = Alignment(horizontal="center")

    header_row = 4

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # -----------------------
    # DATA ROWS
    # -----------------------
    row_idx = header_row + 1

    for item in items:
        ws.cell(row=row_idx, column=1, value=item["item_name"])
        ws.cell(row=row_idx, column=2, value=item["stock"])

        col_offset = 3
        for i, team in enumerate(teams):
            team_data = item["teams"].get(team, {})
            qty = int(team_data.get("requested", 0) or 0)
            ws.cell(row=row_idx, column=col_offset + i, value=qty)

        # remarks column (blank for writing)
        ws.cell(row=row_idx, column=col_offset + len(teams), value="")

        row_idx += 1

    # -----------------------
    # COLUMN WIDTHS
    # -----------------------
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10

    for i in range(len(teams)):
        col_letter = chr(ord("C") + i)
        ws.column_dimensions[col_letter].width = 10

    last_col_letter = chr(ord("C") + len(teams))
    ws.column_dimensions[last_col_letter].width = 20  # Remarks

    # -----------------------
    # FREEZE HEADER
    # -----------------------
    ws.freeze_panes = "A5"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    last_row = row_idx - 1
    last_col = len(headers)

    for row in ws.iter_rows(
        min_row=header_row,
        max_row=last_row,
        min_col=1,
        max_col=last_col
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # item name left aligned
    for r in range(header_row + 1, last_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

    grey_fill = PatternFill("solid", fgColor="D9D9D9")

    for col in range(3, last_col):  # team columns only, excludes Remarks
        if col % 2 == 1:
            for r in range(header_row, last_row + 1):
                ws.cell(row=r, column=col).fill = grey_fill

    # -----------------------
    # EXPORT
    # -----------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def page_packing_fulfillment():
    require_login()
    user = st.session_state["user"]

    if user["role"] not in ("STORE", "ADMIN"):
        st.error("Access denied.")
        return

    st.title("Packing / Fulfillment")
    st.caption("Outstanding items across selected order type.")

    mode_label = st.selectbox(
        "Packing Mode",
        ["Tuesday", "Friday", "Annex - Tuesday", "Annex - Friday", "OT"],
        key="packing_mode"
    )

    mode_map = {
        "Tuesday": "Tuesday",
        "Friday": "Friday",
        "Annex - Tuesday": "ANNEX_TUE",
        "Annex - Friday": "ANNEX_FRI",
        "OT": "OT",
    }

    mode = mode_map[mode_label]

    data = get_packing_list_data(mode)
    items = data["items"]
    teams = data["team_order"]

    if not items:
        st.info("No outstanding items.")

        st.divider()

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Back", use_container_width=True):
                st.session_state["page"] = "home"
                st.rerun()
        with c2:
            if st.button("Orders", use_container_width=True):
                st.session_state["page"] = "orders"
                st.rerun()

        return

    # -----------------------
    # Build editable table rows
    # -----------------------
    table_rows = []
    for item in items:
        row = {
            "Item": item["item_name"],
            "Stock": item["stock"],
            "Status": "⚠ LOW" if item.get("low_stock") else "OK",
        }

        for team in teams:
            team_data = item["teams"].get(team, {})
            req = int(team_data.get("requested", 0) or 0)
            iss = int(team_data.get("issued", 0) or 0)

            if req > 0 and iss >= req:
                stat = "🟢"
            elif iss > 0:
                stat = "🟡"
            else:
                stat = "⚪"

            row[f"{team} Req"] = req
            row[f"{team} Iss"] = iss
            row[f"{team} Stat"] = stat

        table_rows.append(row)

    # -----------------------
    # Editable grid
    # -----------------------
    data_key = f"packing_data_{mode}"
    editor_key = f"packing_editor_{mode}"

    # initialize backing data only once per mode
    if data_key not in st.session_state:
        st.session_state[data_key] = table_rows

    # -----------------------
    # Fulfill Team buttons
    # -----------------------
    if teams:
        st.caption("Quick fill helpers")

        cols = st.columns(len(teams))
        for idx, team in enumerate(teams):
            with cols[idx]:
                if st.button(f"Fulfill {team}", use_container_width=True, key=f"fulfill_{team}_{mode}"):
                    current_rows = st.session_state.get(data_key, table_rows)

                    updated_rows = []
                    for row in current_rows:
                        req_col = f"{team} Req"
                        iss_col = f"{team} Iss"

                        req_val = int(row.get(req_col, 0) or 0)
                        iss_val = int(row.get(iss_col, 0) or 0)

                        # only fill untouched cells
                        if req_val > 0 and iss_val == 0:
                            row[iss_col] = req_val

                        updated_rows.append(row)

                    st.session_state[data_key] = updated_rows
                    st.rerun()

    # -----------------------
    # Editable grid
    # -----------------------
    editor_data = st.session_state.get(data_key, table_rows)

    disabled_cols = ["Item", "Stock", "Status"]
    for team in teams:
        disabled_cols.append(f"{team} Req")
        disabled_cols.append(f"{team} Stat")

    edited = st.data_editor(
        editor_data,
        hide_index=True,
        use_container_width=True,
        disabled=disabled_cols,
        key=editor_key,
    )

    # keep latest edited table rows separately from widget state
    st.session_state[data_key] = edited

    st.divider()

    if st.button("Issue / Save Issued", use_container_width=True):
        try:
            save_packing_board_issued(
                mode=mode,
                edited_rows=edited,
                updated_by=user["username"],
            )
            st.success("Packing board issued quantities saved.")
            if data_key in st.session_state:
                del st.session_state[data_key]
            st.rerun()
        except Exception as e:
            st.error(f"Could not save issued quantities: {e}")

    st.info("This screen is currently prefill + edit only. Save/Issue from this board comes next.")

    st.divider()

    if st.button("Generate Packing List (A3)", use_container_width=True):
        st.session_state["packing_output"] = generate_packing_list_excel()
        st.rerun()

    if "packing_output" in st.session_state:
        st.download_button(
            label="Download Packing List",
            data=st.session_state["packing_output"],
            file_name="packing_list.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Back", use_container_width=True, key="packing_back"):
            st.session_state["page"] = "home"
            st.rerun()
    with c2:
        if st.button("Orders", use_container_width=True, key="packing_orders"):
            st.session_state["page"] = "orders"
            st.rerun()


def page_system_tools():
    require_login()
    user = st.session_state["user"]

    if user["role"] != "ADMIN":
        st.error("Access denied.")
        return

    st.title("System Tools")
    st.caption("Danger zone. Use only during UAT / controlled resets.")

    confirm_text = st.text_input("Type RESET to enable reset actions")

    can_reset = confirm_text.strip().upper() == "RESET"

    st.divider()
    st.subheader("Reset Options")

    c1, c2 = st.columns(2)

    with c1:
        if st.button("Reset Orders Only", use_container_width=True, disabled=not can_reset):
            try:
                reset_orders_only()
                st.success("Orders and order lines reset successfully.")
            except Exception as e:
                st.error(f"Reset failed: {e}")

    with c2:
        if st.button("Reset Orders + Movements", use_container_width=True, disabled=not can_reset):
            try:
                reset_orders_and_movements()
                st.success("Orders, order lines, and stock movements reset successfully.")
            except Exception as e:
                st.error(f"Reset failed: {e}")

    st.divider()

    st.warning(
        "Full database reset is still best done manually by deleting "
        "`data/housekeeping_hub.db` while the app is stopped."
    )

    st.subheader("Find Stock Movement")

    search_item = st.text_input("Item Name", key="movement_search_item")

    date_from = st.date_input("From Date", value=None, key="movement_date_from")
    date_to = st.date_input("To Date", value=None, key="movement_date_to")

    if st.button("Search Movements", use_container_width=True):
        results = search_stock_movements(
            item_name=search_item,
            date_from=date_from,
            date_to=date_to
        )

        if not results:
            st.info("No movements found.")
        else:
            st.dataframe(results, use_container_width=True)

    st.divider()
    st.subheader("Stock Movement Corrections")

    movement_id = st.number_input(
        "Movement ID",
        min_value=1,
        step=1,
        key="void_movement_id"
    )

    void_reason = st.text_input(
        "Void Reason",
        key="void_reason"
    )

    if st.button("Void Stock Movement", use_container_width=True):

        if not void_reason.strip():
            st.warning("Please enter a void reason.")

        else:
            rows_updated = void_stock_movement(
                int(movement_id),
                user["username"],
                void_reason.strip()
            )

            if rows_updated > 0:
                st.success("Movement voided successfully.")
                st.rerun()
            else:
                st.warning(
                    "Movement not found or already voided."
                )

    st.divider()
    st.subheader("Edit Monthly Opening Balance")

    inventory_rows = get_inventory_rows()
    item_options = [r["item_name"] for r in inventory_rows]

    selected_item = st.selectbox(
        "Item",
        item_options,
        key="opening_balance_item"
    )

    opening_month = st.date_input(
        "Month",
        value=date.today().replace(day=1),
        key="opening_balance_month"
    )

    opening_balance = st.number_input(
        "Opening Balance",
        min_value=0,
        step=1,
        key="opening_balance_value"
    )

    if st.button("Save Opening Balance", use_container_width=True):
        month_key = opening_month.strftime("%Y-%m")

        save_opening_balance_override(
            selected_item,
            month_key,
            int(opening_balance)
        )

        st.success("Opening balance saved.")

    st.divider()
    st.subheader("Edit Stock Movement Date")

    edit_movement_id = st.number_input(
        "Movement ID to Edit",
        min_value=1,
        step=1,
        key="edit_movement_date_id"
    )

    new_movement_date = st.date_input(
        "New Movement Date",
        value=date.today(),
        key="edit_movement_new_date"
    )

    if st.button("Update Movement Date", use_container_width=True):
        rows_updated = update_stock_movement_date(
            int(edit_movement_id),
            str(new_movement_date)
        )

        if rows_updated > 0:
            st.success("Movement date updated.")
            st.rerun()
        else:
            st.warning("Movement not found or already voided.")

    st.divider()
    st.subheader("Linen UAT Tools")

    linen_cycles = get_linen_cycles()

    if not linen_cycles:
        st.info("No linen inventory cycles found.")
    else:
        cycle_options = {
            f"{c['id']} - {c['cycle_name']} ({c['status']})": c["id"]
            for c in linen_cycles
        }

        selected_cycle_label = st.selectbox(
            "Linen Cycle",
            list(cycle_options.keys()),
            key="uat_linen_cycle_select"
        )

        if st.button(
            "Force Complete Linen Inventory",
            use_container_width=True
        ):
            force_complete_linen_cycle(
                cycle_options[selected_cycle_label]
            )

            st.success("Linen inventory force-completed for UAT.")
            st.rerun()

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "home"
        st.rerun()

def page_glo_gel_audit_detail():
    require_login()
    user = st.session_state["user"]

    audit_id = st.session_state.get("selected_audit_id")
    if not audit_id:
        st.warning("No audit selected.")
        return

    audit = get_audit_header(int(audit_id))
    if not audit:
        st.error("Audit not found.")
        return

    st.title("Glo Gel Audit Detail")

    is_completed = audit["status"] == "COMPLETED"

    st.caption(
        f"{audit['location_name']} | {audit['audit_date']} | "
        f"Status: {audit['status']} | Created by {audit['created_by']}"
    )

    audit_date = st.date_input(
        "Date of Inspection",
        value=date.fromisoformat(audit["audit_date"]),
        disabled=is_completed,
        key=f"audit_date_{audit_id}"
    )
    auditor_name = st.text_input(
        "Auditor Name",
        value=audit["auditor_name"],
        disabled=is_completed,
        key=f"audit_auditor_{audit_id}"
    )
    staff_name = st.text_input(
        "Staff Name",
        value=audit["staff_name"],
        disabled=is_completed,
        key=f"audit_staff_{audit_id}"
    )

    remarks = st.text_area(
        "Remarks",
        value=audit["remarks"] or "",
        disabled=is_completed,
        key=f"audit_remarks_{audit_id}"
    )

    surfaces = get_surfaces_for_template(audit["template_group"])
    saved_results = get_audit_results(int(audit_id))

    saved_standard = {
        r["surface_id"]: r["result"]
        for r in saved_results
        if r["is_additional"] == "N"
    }
    saved_additional = [
        r for r in saved_results if r["is_additional"] == "Y"
    ]

    st.divider()
    st.subheader("High Touch Surfaces")

    grouped = {}
    for s in surfaces:
        area = s["area_group"] or "General"
        grouped.setdefault(area, []).append(s)

    if audit["template_group"] == "B":
        area_order = ["Patient", "Toilet", "General"]
    else:
        area_order = ["Clinic", "Toilet", "General"]

    ordered_areas = (
        [a for a in area_order if a in grouped] +
        [a for a in grouped.keys() if a not in area_order]
    )

    standard_results = []

    for area in ordered_areas:
        items = grouped[area]
        st.markdown(f"**{area}**")

        for s in items:
            default_result = saved_standard.get(s["surface_id"], "NA")
            choice = st.radio(
                s["surface_name"],
                options=["C", "NC", "NA"],
                horizontal=True,
                disabled=is_completed,
                index=["C", "NC", "NA"].index(default_result),
                key=f"audit_{audit_id}_{s['surface_id']}"
            )

            standard_results.append({
                "surface_id": s["surface_id"],
                "surface_name": s["surface_name"],
                "result": choice,
                "area_group": s["area_group"],
                "display_order": s["display_order"],
            })

    st.divider()
    st.subheader("Additional HTS")

    additional_results = []

    for i in range(3):
        existing_name = saved_additional[i]["surface_name"] if i < len(saved_additional) else ""
        existing_result = saved_additional[i]["result"] if i < len(saved_additional) else "NA"

        c1, c2 = st.columns([3, 2])

        with c1:
            name = st.text_input(
                f"Surface {i+1}",
                value=existing_name,
                disabled=is_completed,
                key=f"add_name_{audit_id}_{i}"
            )

        with c2:
            result = st.radio(
                "",
                options=["C", "NC", "NA"],
                horizontal=True,
                disabled=is_completed,
                index=["C", "NC", "NA"].index(existing_result),
                key=f"add_result_{audit_id}_{i}",
                label_visibility="collapsed"
            )

        additional_results.append({
            "surface_name": name,
            "result": result,
        })

    if not is_completed:
        c1, c2 = st.columns(2)

        with c1:
            if st.button("Save Draft", use_container_width=True, key=f"save_draft_{audit_id}"):
                save_audit_detail(
                    audit_id=int(audit_id),
                    audit_date=str(audit_date),
                    auditor_name=auditor_name,
                    staff_name=staff_name,
                    remarks=remarks,
                    standard_results=standard_results,
                    additional_results=additional_results,
                    final_status="DRAFT",
                )
                st.session_state["audit_success"] = True
                st.rerun()

        with c2:
            if st.button("Complete Audit", use_container_width=True, key=f"complete_audit_{audit_id}"):
                if not staff_name.strip():
                    st.warning("Please enter staff name.")
                else:
                    save_audit_detail(
                        audit_id=int(audit_id),
                        audit_date=str(audit_date),
                        auditor_name=auditor_name,
                        staff_name=staff_name,
                        remarks=remarks,
                        standard_results=standard_results,
                        additional_results=additional_results,
                        final_status="COMPLETED",
                    )
                    st.session_state["audit_success"] = True
                    st.session_state["page"] = "glo_gel_audits"
                    st.rerun()

    st.divider()
    if st.button("Back", use_container_width=True, key=f"audit_detail_back_{audit_id}"):
        st.session_state["page"] = "glo_gel_audits"
        st.rerun()
        
def page_glo_gel_audits():
    require_login()
    user = st.session_state["user"]

    st.title("Glo Gel Audits")
    
    glo_from = st.date_input(
        "Report From Date",
        value=date.today().replace(day=1),
        key="glo_report_from"
    )

    glo_to = st.date_input(
        "Report To Date",
        value=date.today(),
        key="glo_report_to"
    )

    # ---------------------------
    # Success message
    # ---------------------------
    if st.session_state.pop("audit_success", False):
        st.success("Audit saved successfully.")

    if user["role"] == "ADMIN":
        st.divider()
        st.subheader("Admin Report Export")

        if st.button("Prepare Compiled Audit Report", use_container_width=True, key="prepare_compiled_audit_report"):
            try:
                st.session_state["compiled_audit_report"] = generate_compiled_glo_gel_report_excel(date_from=str(glo_from),date_to=str(glo_to))
                st.rerun()
            except Exception as e:
                st.error(str(e))

        if "compiled_audit_report" in st.session_state:
            st.download_button(
                label="Download Compiled Audit Report",
                data=st.session_state["compiled_audit_report"],
                file_name="compiled_glo_gel_audit_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_compiled_audit_report"
            )

    # ---------------------------
    # Create New Audit
    # ---------------------------
    st.subheader("Create New Audit")

    zones = get_visible_zones_for_user(user)

    if not zones:
        st.warning("No zones available for your group.")
    else:
        selected_zone = st.selectbox(
            "Zone",
            zones,
            key="new_audit_zone"
        )

        locations = get_visible_locations_for_user_and_zone(user, selected_zone)

        if not locations:
            st.warning("No locations available for the selected zone.")
        else:
            location_names = [l["location_name"] for l in locations]
            selected_location_name = st.selectbox(
                "Area / Location",
                location_names,
                key="new_audit_location"
            )

            selected_location = next(
                l for l in locations if l["location_name"] == selected_location_name
            )

        auditor_name = st.text_input(
            "Auditor Name",
            value=user["username"],
            key="new_audit_auditor"
        )

        staff_name = st.text_input(
            "Staff Name",
            key="new_audit_staff"
        )

        audit_date = st.date_input(
            "Date of Inspection",
            key="new_audit_date"
        )

        remarks = st.text_area(
            "Remarks",
            key="new_audit_remarks"
        )

        if st.button("Create Audit Entry", use_container_width=True):
            if not staff_name.strip():
                st.warning("Please enter staff name.")
            else:
                template_group = get_template_for_tower(selected_location["tower"])

                audit_id = create_audit_header(
                    audit_date=str(audit_date),
                    auditor_name=auditor_name,
                    staff_name=staff_name,
                    location=selected_location,
                    template_group=template_group,
                    remarks=remarks,
                    created_by=user["username"],
                )

                if not audit_id:
                    st.error("Audit was not created. No audit_id returned.")
                    return

                st.session_state["selected_audit_id"] = audit_id
                st.session_state["page"] = "glo_gel_audit_detail"
                st.rerun()

    # ---------------------------
    # Audit History
    # ---------------------------
    st.divider()
    st.subheader("Audit History")

    audits = list_audits_for_user(user)

    if not audits:
        st.info("No audits yet.")
    else:
        # Filters
        col1, col2, col3 = st.columns(3)

        with col1:
            filter_status = st.selectbox(
                "Status",
                ["All", "DRAFT", "COMPLETED"],
                key="audit_filter_status"
            )

        with col2:
            filter_location = st.selectbox(
                "Location",
                ["All"] + sorted({a["location_name"] for a in audits}),
                key="audit_filter_location"
            )

        with col3:
            filter_staff = st.text_input(
                "Search Staff",
                key="audit_filter_staff"
            )

        # Apply filters
        filtered = []

        for a in audits:
            if filter_status != "All" and a["status"] != filter_status:
                continue

            if filter_location != "All" and a["location_name"] != filter_location:
                continue

            if filter_staff and filter_staff.lower() not in (a["staff_name"] or "").lower():
                continue

            filtered.append(a)

        # Display list
        for a in filtered:
            c1, c2 = st.columns([6, 1])

            status_icon = "🟡" if a["status"] == "DRAFT" else "🟢"

            with c1:
                status_icon = "🟡" if a["status"] == "DRAFT" else "🟢"

                zone = (a.get("zone") or "").strip()
                location_display = f"{zone} - {a['location_name']}" if zone else a["location_name"]

                st.write(
                    f"📅 {a['audit_date']} | "
                    f"📍 {location_display} | "
                    f"👤 {a['staff_name']} | "
                    f"🧑‍⚕️ {a['auditor_name']} | "
                    f"{status_icon} {a['status']}"
            )

            with c2:
                if st.button(
                    "Open",
                    key=f"history_open_audit_{a['audit_id']}",  # <- UNIQUE KEY FIX
                    use_container_width=True
                ):
                    st.session_state["selected_audit_id"] = a["audit_id"]
                    st.session_state["page"] = "glo_gel_audit_detail"
                    st.rerun()

    # ---------------------------
    # Back button
    # ---------------------------
    st.divider()

    if st.button("Back", use_container_width=True, key="audits_back"):
        st.session_state["page"] = "home"
        st.rerun()

def generate_audit_export_excel(audit_id: int):
    audit = get_audit_header(audit_id)
    if not audit:
        raise ValueError("Audit not found.")

    results = get_audit_results(audit_id)

    standard_results = [r for r in results if r["is_additional"] == "N"]
    additional_results = [r for r in results if r["is_additional"] == "Y"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"

    # ---------------------------
    # Styles
    # ---------------------------
    bold = Font(bold=True)
    title_font = Font(size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------------------------
    # Header
    # ---------------------------
    ws["A1"] = "Glo Gel Audit"
    ws["A1"].font = title_font

    ws["A3"] = "Auditor Name"
    ws["B3"] = audit["auditor_name"]

    ws["D3"] = "Staff Name"
    ws["E3"] = audit["staff_name"]

    ws["A4"] = "Date of Inspection"
    ws["B4"] = audit["audit_date"]

    ws["D4"] = "Location"
    ws["E4"] = audit["location_name"]

    ws["A5"] = "Template"
    ws["B5"] = audit["template_group"]

    for cell in ["A3", "D3", "A4", "D4", "A5"]:
        ws[cell].font = bold

    # ---------------------------
    # Surface Results Table
    # ---------------------------
    start_row = 7
    ws[f"A{start_row}"] = "Surface"
    ws[f"B{start_row}"] = "C"
    ws[f"C{start_row}"] = "NC"
    ws[f"D{start_row}"] = "NA"

    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{start_row}"].font = bold
        ws[f"{col}{start_row}"].alignment = center
        ws[f"{col}{start_row}"].border = border

    current_row = start_row + 1

    # Group standard surfaces by area_group
    grouped = {}
    for r in standard_results:
        area = r["area_group"] or "General"
        grouped.setdefault(area, []).append(r)

    for area, items in grouped.items():
        ws[f"A{current_row}"] = area
        ws[f"A{current_row}"].font = bold
        current_row += 1

        items.sort(key=lambda x: (x["display_order"] or 0, x["surface_name"]))

        for r in items:
            ws[f"A{current_row}"] = r["surface_name"]

            result = r["result"]
            ws[f"B{current_row}"] = "C" if result == "C" else ""
            ws[f"C{current_row}"] = "NC" if result == "NC" else ""
            ws[f"D{current_row}"] = "NA" if result == "NA" else ""

            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{current_row}"].border = border

            ws[f"B{current_row}"].alignment = center
            ws[f"C{current_row}"].alignment = center
            ws[f"D{current_row}"].alignment = center

            current_row += 1

        current_row += 1

    # ---------------------------
    # Additional HTS
    # ---------------------------
    ws[f"A{current_row}"] = "Additional HTS"
    ws[f"A{current_row}"].font = bold
    current_row += 1

    ws[f"A{current_row}"] = "Surface"
    ws[f"B{current_row}"] = "C"
    ws[f"C{current_row}"] = "NC"
    ws[f"D{current_row}"] = "NA"

    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{current_row}"].font = bold
        ws[f"{col}{current_row}"].alignment = center
        ws[f"{col}{current_row}"].border = border

    current_row += 1

    if additional_results:
        for r in additional_results:
            ws[f"A{current_row}"] = r["surface_name"]

            result = r["result"]
            ws[f"B{current_row}"] = "C" if result == "C" else ""
            ws[f"C{current_row}"] = "NC" if result == "NC" else ""
            ws[f"D{current_row}"] = "NA" if result == "NA" else ""

            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{current_row}"].border = border

            ws[f"B{current_row}"].alignment = center
            ws[f"C{current_row}"].alignment = center
            ws[f"D{current_row}"].alignment = center

            current_row += 1
    else:
        ws[f"A{current_row}"] = "-"
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{current_row}"].border = border
        current_row += 1

    current_row += 1

    # ---------------------------
    # Remarks
    # ---------------------------
    ws[f"A{current_row}"] = "Remarks"
    ws[f"A{current_row}"].font = bold
    ws[f"B{current_row}"] = audit["remarks"] or ""

    # ---------------------------
    # Column widths
    # ---------------------------
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_compiled_glo_gel_report_excel(date_from=None, date_to=None):
    grouped = get_completed_audits_grouped_by_tower(date_from, date_to)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    bold = Font(bold=True)
    title_font = Font(size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet_specs = [
        ("A", "Tower A"),
        ("B", "Tower B"),
        ("C", "Tower C"),
    ]

    def write_audit_block(ws, start_row, start_col, audit_entry, template_surfaces):
        audit = audit_entry["audit"]
        standard_results = audit_entry["standard_results"]
        additional_results = audit_entry["additional_results"]

        result_lookup = {
            r["surface_id"]: r["result"]
            for r in standard_results
            if r.get("surface_id")
        }

        # Column map for one block
        c0 = start_col       # label/value col 1
        c1 = start_col + 1   # label/value col 2
        c2 = start_col + 2   # result C
        c3 = start_col + 3   # result NC
        c4 = start_col + 4   # result NA

        # Header rows
        ws.cell(start_row, c0, "Date").font = bold
        ws.cell(start_row, c1, audit["audit_date"])

        ws.cell(start_row + 1, c0, "Level").font = bold
        ws.cell(start_row + 1, c1, audit.get("zone") or "")

        ws.cell(start_row + 2, c0, "Location").font = bold
        ws.cell(start_row + 2, c1, audit["location_name"])

        ws.cell(start_row + 3, c0, "Staff").font = bold
        ws.cell(start_row + 3, c1, audit["staff_name"])

        ws.cell(start_row + 4, c0, "Auditor").font = bold
        ws.cell(start_row + 4, c1, audit["auditor_name"])

        ws.cell(start_row + 5, c0, "Remarks").font = bold
        ws.cell(start_row + 5, c1, audit.get("remarks") or "")

        # Surface header
        header_row = start_row + 7
        ws.merge_cells(start_row=header_row, start_column=c0, end_row=header_row, end_column=c1)
        ws.merge_cells(start_row=header_row, start_column=c0, end_row=header_row, end_column=c1)

        for col in range(c0, c1 + 1):
            ws.cell(header_row, col).border = border

        ws.cell(header_row, c0).value = "Surface"
        ws.cell(header_row, c0).font = bold
        ws.cell(header_row, c0).alignment = center
        ws.cell(header_row, c2, "C").font = bold
        ws.cell(header_row, c3, "NC").font = bold
        ws.cell(header_row, c4, "NA").font = bold

        for col in [c0, c2, c3, c4]:
            ws.cell(header_row, col).alignment = center
            ws.cell(header_row, col).border = border

        # Surface rows
        row = header_row + 1
        for s in template_surfaces:
            ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)

            for col in range(c0, c1 + 1):
                cell = ws.cell(row, col)
                cell.border = border

            ws.cell(row, c0).value = s["surface_name"]
            ws.cell(row, c0).alignment = Alignment(vertical="center")

            result = result_lookup.get(s["surface_id"], "")
            ws.cell(row, c2, "C" if result == "C" else "")
            ws.cell(row, c3, "NC" if result == "NC" else "")
            ws.cell(row, c4, "NA" if result == "NA" else "")

            for col in [c2, c3, c4]:
                ws.cell(row, col).alignment = center
                ws.cell(row, col).border = border

            row += 1

        # Additional HTS
        # Additional HTS title row
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)

        for col in range(c0, c1 + 1):
            ws.cell(row, col).border = border

        ws.cell(row, c0).value = "Additional HTS"
        ws.cell(row, c0).font = bold

        row += 1

        # Clear entire row first (important)
        for col in range(c0, c4 + 1):
            ws.cell(row, col).border = Border()

        # Now apply only the borders you want
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)

        for col in range(c0, c1 + 1):
            ws.cell(row, col).border = border

        ws.cell(row, c0).value = "-"

        for col in [c2, c3, c4]:
            ws.cell(row, col).border = border

        return row

    for tower_code, sheet_name in sheet_specs:
        ws = wb.create_sheet(title=sheet_name)
        audits = grouped.get(tower_code, [])
        template_surfaces = get_surface_template_for_tower(tower_code)

        ws["A1"] = f"Glo Gel Audit Report - {sheet_name}"
        ws["A1"].font = title_font

        if not audits:
            ws["A3"] = "No completed audits."
            continue

        # Layout:
        # left block starts at col 1 (A)
        # right block starts at col 8 (H)
        left_start_col = 1
        right_start_col = 8

        # block height depends on surface count
        # 5 header rows + blank + surface header + surfaces + add hts title + at least 1 row + gap
        block_height = 6 + 1 + len(template_surfaces) + 1 + 1 + 2

        for idx, entry in enumerate(audits):
            pair_index = idx // 2
            is_right = (idx % 2 == 1)

            start_row = 3 + pair_index * block_height
            start_col = right_start_col if is_right else left_start_col

            write_audit_block(ws, start_row, start_col, entry, template_surfaces)

        # widths for left block
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8

        # spacer columns
        ws.column_dimensions["F"].width = 4
        ws.column_dimensions["G"].width = 4

        # widths for right block
        ws.column_dimensions["H"].width = 18
        ws.column_dimensions["I"].width = 22
        ws.column_dimensions["J"].width = 8
        ws.column_dimensions["K"].width = 8
        ws.column_dimensions["L"].width = 8

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def page_linen_manage():
    require_login()
    user = st.session_state["user"]
    role = user["role"]

    if role not in ("ADMIN", "LINSUP"):
        st.error("Access denied.")
        return

    st.title("Manage Linen Inventory")

    # ADMIN + LINSUP only
    if role in ("ADMIN", "LINSUP"):
        st.subheader("Create New Inventory")

        cycle_name = st.text_input(
            "Inventory Name",
            value="Linen Inventory",
            key="manage_linen_cycle_name"
        )

        if st.button(
            "Create Linen Inventory",
            use_container_width=True,
            key="manage_create_linen_inventory"
        ):
            if not cycle_name.strip():
                st.warning("Please enter an inventory name.")
            else:
                create_linen_cycle(
                    cycle_name.strip(),
                    user["username"]
                )
                st.success("Inventory created.")
                st.rerun()

    st.divider()
    st.subheader("Inventory Cycles")

    cycles = get_linen_cycles()

    if not cycles:
        st.info("No linen inventory cycles found.")
    else:
        for cycle in cycles:
            st.markdown(f"### {cycle['cycle_name']}")
            st.write(f"Status: `{cycle['status']}`")
            st.write(f"Created by: **{cycle['created_by']}**")
            st.write(f"Created at: {cycle['created_at']}")

            if st.button(
                "Open",
                key=f"manage_open_linen_cycle_{cycle['id']}",
                use_container_width=True
            ):
                st.session_state["active_linen_cycle_id"] = cycle["id"]
                st.session_state["page"] = "linen_cycle_detail"
                st.rerun()

            st.divider()

    if st.button("Back", use_container_width=True, key="manage_linen_back"):
        st.session_state["page"] = "home"
        st.rerun()


def page_linen_inventory():
    require_login()
    user = st.session_state["user"]

    role = user["role"]

    active_cycle = get_active_linen_cycle()

    if role.startswith("LINREP"):

        st.subheader("Active Inventory")

        if not active_cycle:
            st.info("No active linen inventory.")
            return

        st.success(
            f"Active Cycle: {active_cycle['cycle_name']}"
        )

        assignments = get_assignments_for_user(
            active_cycle["id"],
            user["username"]
        )

        status_map = get_submission_status_map(active_cycle["id"])

        if not assignments:
            st.info(
                "No locations assigned."
            )
            return

        st.subheader("Assigned Locations")

        location_map = get_linen_location_map()
        seen_location_ids = set()

        for row in assignments:
            location_id = row["location_id"]

            if location_id in seen_location_ids:
                continue

            seen_location_ids.add(location_id)

            loc = location_map.get(location_id)

            if not loc:
                st.write(location_id)
                continue

            location_label = (
                f"{loc['tower']} - {loc['level']} - "
                f"{loc['zone']} - {loc['location_name']}"
            )

            status = status_map.get(location_id, "PENDING")

            col1, col2, col3 = st.columns([4, 1, 1])

            with col1:
                st.write(location_label)

            with col2:
                st.write(status)

            with col3:
                if st.button("Open", key=f"open_linen_count_{location_id}", use_container_width=True):
                    st.session_state["linen_count_location_id"] = location_id
                    st.session_state["page"] = "linen_count"
                    st.rerun()

        return

    if role == "TEAM":

        active_cycle = get_active_linen_cycle()

        if not active_cycle:
            st.info("No active linen inventory.")
            return

        location_rows = load_linen_locations_rows()

        group_column_map = {
            "B1-4": "lin_B1-4",
            "B5-10": "lin_B5-10",
            "B11-16": "lin_B11-16",
            "C1-12": "lin_C1-12",
        }

        access_column = group_column_map.get(
            user["group"]
        )

        allowed_locations = [
            row
            for row in location_rows
            if str(
                row.get(access_column, "")
            ).strip().upper() == "Y"
        ]

        st.title("Linen Inventory")

        st.subheader("Active Inventory")
        st.success(f"Active Cycle: {active_cycle['cycle_name']}")

        status_map = get_submission_status_map(active_cycle["id"])

        st.subheader("Assigned Locations")

        if not allowed_locations:
            st.info("No linen locations assigned to your team.")
            return

        for loc in allowed_locations:
            location_id = loc["location_id"]

            location_label = (
                f"{loc['tower']} - {loc['level']} - "
                f"{loc['zone']} - {loc['location_name']}"
            )

            status = status_map.get(location_id, "PENDING")

            col1, col2, col3 = st.columns([4, 1, 1])

            with col1:
                st.write(location_label)

            with col2:
                st.write(status)

            with col3:
                if st.button(
                    "Open",
                    key=f"open_team_linen_count_{location_id}",
                    use_container_width=True
                ):
                    st.session_state["linen_count_location_id"] = location_id
                    st.session_state["page"] = "linen_count"
                    st.rerun()

        return

    if role in ("LINTEAM", "LINSUP", "ADMIN"):

        st.title("Linen Inventory")

        if active_cycle:
            st.subheader("Active Inventory")
            st.success(f"Active Cycle: {active_cycle['cycle_name']}")

            locations = load_linen_locations_rows()
            status_map = get_submission_status_map(active_cycle["id"])

            st.subheader("All Locations")

            for loc in locations:
                location_id = loc["location_id"]

                location_label = (
                    f"{loc['tower']} - {loc['level']} - "
                    f"{loc['zone']} - {loc['location_name']}"
                )

                status = status_map.get(location_id, "PENDING")

                col1, col2, col3 = st.columns([4, 1, 1])

                with col1:
                    st.write(location_label)

                with col2:
                    st.write(status)

                with col3:
                    if st.button(
                        "Open",
                        key=f"open_all_linen_count_{location_id}",
                        use_container_width=True
                    ):
                        st.session_state["linen_count_location_id"] = location_id
                        st.session_state["page"] = "linen_count"
                        st.rerun()

            return        

    st.title("Linen Inventory")

    st.subheader("Inventory Cycles")

    cycles = get_linen_cycles()

    if not cycles:
        st.info("No linen inventory cycles found.")
    else:

        for cycle in cycles:
            st.markdown(f"### {cycle['cycle_name']}")
            st.write(f"Status: `{cycle['status']}`")
            st.write(f"Created by: **{cycle['created_by']}**")
            st.write(f"Created at: {cycle['created_at']}")

            if st.button(
                "Open",
                key=f"open_linen_cycle_{cycle['id']}",
                use_container_width=True
            ):
                st.session_state["active_linen_cycle_id"] = cycle["id"]
                st.session_state["page"] = "linen_cycle_detail"
                st.rerun()

            st.divider()

def page_linen_cycle_detail():
    require_login()

    cycle_id = st.session_state.get("active_linen_cycle_id")

    if not cycle_id:
        st.warning("No linen inventory selected.")
        st.session_state["page"] = "linen_inventory"
        st.rerun()

    cycle = get_linen_cycle(cycle_id)

    saved_reps = get_cycle_reps(cycle_id)

    saved_assignments = get_cycle_assignments(cycle_id)

    assigned_map = {}
    for row in saved_assignments:
        assigned_map.setdefault(row["assigned_to"], []).append(row["location_id"])

    saved_rep_map = {
        row["rep_username"]: row["display_name"]
        for row in saved_reps
    }

    if not cycle:
        st.error("Linen inventory not found.")
        st.session_state["page"] = "linen_inventory"
        st.rerun()

    st.title(cycle["cycle_name"])
    st.write(f"Status: `{cycle['status']}`")
    st.write(f"Created by: **{cycle['created_by']}**")
    st.write(f"Created at: {cycle['created_at']}")

    st.divider()

    st.subheader("Setup")
    st.subheader("Linen Representatives")

    for i in range(1, 11):
        rep_username = f"LINREP{i}"
        saved_name = saved_rep_map.get(rep_username, "")

        if f"rep_enabled_{i}" not in st.session_state:
            st.session_state[f"rep_enabled_{i}"] = bool(saved_name)

        if f"rep_name_{i}" not in st.session_state:
            st.session_state[f"rep_name_{i}"] = saved_name

        col1, col2 = st.columns([1, 4])

        with col1:
            enabled = st.checkbox(
                f"Linen Rep {i}",
                key=f"rep_enabled_{i}"
            )

        with col2:
            st.text_input(
                "Name",
                key=f"rep_name_{i}",
                disabled=not enabled,
                label_visibility="collapsed",
                placeholder="Enter name..."
            )

    if st.button("Save Representatives", use_container_width=True):
        rep_rows = []

        for i in range(1, 11):
            enabled = st.session_state.get(f"rep_enabled_{i}", False)
            name = st.session_state.get(f"rep_name_{i}", "").strip()

            if enabled:
                if not name:
                    st.warning(f"Please enter name for Linen Rep {i}.")
                    st.stop()

                rep_rows.append({
                    "rep_username": f"LINREP{i}",
                    "display_name": name,
                })

        save_cycle_reps(cycle_id, rep_rows)
        st.success("Representatives saved.")
        st.rerun()

    st.divider()
    st.subheader("Assignment Summary")

    locations = load_linen_locations_rows()

    total_locations = len(locations)
    assigned_location_ids = {
        row["location_id"]
        for row in saved_assignments
    }

    assigned_locations = len(assigned_location_ids)
    remaining_locations = total_locations - assigned_locations

    c1, c2, c3 = st.columns(3)

    c1.metric("Total Locations", total_locations)
    c2.metric("Assigned", assigned_locations)
    c3.metric("Remaining", remaining_locations)

    st.divider()
    st.subheader("Location Assignment")

    active_reps = [
        {
            "rep_username": row["rep_username"],
            "display_name": row["display_name"],
        }
        for row in saved_reps
    ]

    location_lookup = {}
    location_options = []

    for loc in locations:
        label = f"{loc['tower']} - {loc['level']} - {loc['zone']} - {loc['location_name']}"
        location_lookup[label] = loc["location_id"]
        location_options.append(label)

    if not active_reps:
        st.info("Save representatives first before assigning locations.")
    else:
        assignment_rows = []

        for rep in active_reps:
            rep_key = rep["rep_username"]
            saved_location_ids = assigned_map.get(rep_key, [])

            default_labels = [
                label for label, loc_id in location_lookup.items()
                if loc_id in saved_location_ids
            ]

            selected_labels = st.multiselect(
                f"{rep['display_name']} ({rep_key})",
                options=location_options,
                default=default_labels,
                key=f"assign_locations_{rep_key}",
            )

            for label in selected_labels:
                assignment_rows.append({
                    "location_id": location_lookup[label],
                    "assigned_to": rep_key,
                    "assigned_type": "REP",
                })

        if st.button("Save Location Assignments", use_container_width=True):
            save_cycle_assignments(cycle_id, assignment_rows)
            st.success("Location assignments saved.")
            st.rerun()

    st.divider()
    st.subheader("Inventory Progress")

    completed_locations = get_submitted_location_count(
        cycle_id
    )

    pending_count = (
        total_locations
        - completed_locations
    )

    c1, c2, c3 = st.columns(3)

    c1.metric("Total Locations", total_locations)
    c2.metric("Completed", completed_locations)
    c3.metric("Pending", pending_count)

    if total_locations > 0:
        progress_value = completed_locations / total_locations
    else:
        progress_value = 0

    st.progress(progress_value)
    st.caption(f"{completed_locations} of {total_locations} locations submitted")

    status_map = get_submission_status_map(cycle_id)

    st.subheader("Rep Progress")

    if not saved_assignments:
        st.info("No locations assigned to Linen Reps.")
    else:
        rep_name_map = {
            row["rep_username"]: row["display_name"]
            for row in saved_reps
        }

        rep_progress = {}

        for row in saved_assignments:
            assigned_to = row["assigned_to"]
            location_id = row["location_id"]

            rep_progress.setdefault(assigned_to, {
                "total": 0,
                "submitted": 0
            })

            rep_progress[assigned_to]["total"] += 1

            if status_map.get(location_id) == "SUBMITTED":
                rep_progress[assigned_to]["submitted"] += 1

        for rep_username, data in rep_progress.items():
            display_name = rep_name_map.get(rep_username, rep_username)

            st.write(
                f"**{display_name} ({rep_username})** — "
                f"{data['submitted']} / {data['total']} submitted"
            )

    st.divider()
    st.subheader("Start Inventory")

    if cycle["status"] == "DRAFT":
        if st.button("Start Linen Inventory", use_container_width=True):
            start_linen_cycle(cycle_id)
            st.success("Linen inventory started.")
            st.rerun()
    else:
        st.info("This inventory has already been started.")

    st.divider()
    st.subheader("Complete Inventory")

    all_locations_done = (
        completed_locations
        >= total_locations
    )
    
    if cycle["status"] == "ACTIVE":

        if not all_locations_done:

            st.warning(
                f"{pending_count} locations still require submission before inventory can be completed."
            )

        else:

            st.success(
                "All locations have been submitted."
            )

            if st.button(
                "Complete Linen Inventory",
                use_container_width=True
            ):
                complete_linen_cycle(cycle_id)

                st.success(
                    "Linen inventory completed."
                )

            st.rerun()

    elif cycle["status"] == "COMPLETED":
        st.success("This linen inventory has been completed.")

    st.divider()
    st.subheader("Report Data Test")

    if cycle["status"] == "COMPLETED":
        st.divider()
        st.subheader("Linen Report Export")

        report_file = generate_linen_report_excel(cycle_id)

        st.download_button(
            label="Download Linen Inventory Report",
            data=report_file,
            file_name=f"linen_inventory_{cycle_id}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "linen_manage"
        st.rerun()

def page_linen_count():
    require_login()

    user = st.session_state["user"]

    location_id = st.session_state.get("linen_count_location_id")

    if not location_id:
        st.warning("No linen location selected.")
        st.session_state["page"] = "linen_inventory"
        st.rerun()

    location_map = get_linen_location_map()
    loc = location_map.get(location_id)

    st.title("Linen Count")

    if loc:
        st.write(
            f"**Location:** {loc['tower']} - {loc['level']} - "
            f"{loc['zone']} - {loc['location_name']}"
        )
    else:
        st.write(f"**Location:** {location_id}")

    active_cycle = get_active_linen_cycle()

    saved_submission = get_submission(
        active_cycle["id"],
        location_id
    )

    is_submitted = False

    if saved_submission:
        is_submitted = (
            saved_submission["status"] == "SUBMITTED"
        )

    saved_qty_map = {}

    if saved_submission:

        saved_lines = get_submission_lines(
            saved_submission["id"]
        )

        saved_qty_map = {
            row["item_no"]: row["quantity"]
            for row in saved_lines
        }

    items = get_linen_items_for_location(location_id)

    if not items:
        st.warning("No linen items configured for this location.")

    else:

        st.subheader("Items to Count")

        count_lines = []

        can_edit = (
            not is_submitted
            or user["role"] in ("LINTEAM", "LINSUP", "ADMIN")
        )

        for item in items:

            qty = st.number_input(
                item["item_name"],
                min_value=0,
                step=1,
                value=int(
                    saved_qty_map.get(
                        item["item_no"],
                        0
                    )
                ),
                disabled=not can_edit,
                key=f"linen_count_{location_id}_{item['item_no']}"
            )

            count_lines.append({
                "item_no": item["item_no"],
                "quantity": qty
            })

    if can_edit:

        col1, col2 = st.columns(2)

        with col1:

            if st.button(
                "Save Draft",
                use_container_width=True,
                key=f"save_linen_draft_{location_id}"
            ):
                save_submission_draft(
                    active_cycle["id"],
                    location_id,
                    user["username"],
                    count_lines
                )

                st.success("Draft saved.")

        with col2:

            if st.button(
                "Submit Count",
                use_container_width=True,
                key=f"submit_linen_count_{location_id}"
            ):

                save_submission_draft(
                    active_cycle["id"],
                    location_id,
                    user["username"],
                    count_lines
                )

                submit_submission(
                    active_cycle["id"],
                    location_id,
                    user["username"]
                )

                st.success("Count submitted.")
                st.rerun()
    else:
        st.info("This count has been submitted and locked.")

    if st.button("Back", use_container_width=True):
        st.session_state["page"] = "linen_inventory"
        st.rerun()

def generate_linen_report_excel(cycle_id):
    report_rows = get_cycle_submission_lines(cycle_id)

    location_map = get_linen_location_map()
    linen_items = get_linen_items_for_location  # not used here

    # Build item lookup from Linen Master
    from master_loader import load_linen_master_rows
    item_rows = load_linen_master_rows()

    item_order = []
    item_map = {}

    for row in item_rows:
        item_no = str(row["item_no"])

        item_order.append(item_no)

        item_map[item_no] = {
            "item_name": row["item_name"],
            "lin_category": row.get("lin_category", "")
        }

    qty_map = {}

    for row in report_rows:

        location_id = row["location_id"]
        item_no = str(row["item_no"])

        qty_map[(location_id, item_no)] = row["quantity"]

    item_map = {
        str(row["item_no"]): row
        for row in item_rows
    }

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    cycle = get_linen_cycle(cycle_id)

    summary_ws["A1"] = "Linen Inventory Report"
    summary_ws["A2"] = f"Inventory: {cycle['cycle_name']}"
    summary_ws["A3"] = f"Status: {cycle['status']}"
    summary_ws["A4"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"

    summary_ws["A1"].font = Font(size=14, bold=True)
    summary_ws["A2"].font = Font(bold=True)
    summary_ws["A3"].font = Font(bold=True)
    summary_ws["A4"].font = Font(bold=True)

    def write_location_matrix_sheet(sheet_name, location_filter):
        ws = wb.create_sheet(title=sheet_name)

        selected_locations = [
            loc
            for loc in location_map.values()
            if location_filter(loc)
        ]

        selected_locations.sort(
            key=lambda x: (
                x.get("tower", ""),
                x.get("level", ""),
                x.get("zone", ""),
                x.get("location_name", "")
            )
        )

        # headers
        ws.cell(1, 1, "Item No")
        ws.cell(1, 2, "Item Name")

        for col_idx, loc in enumerate(selected_locations, start=3):
            ws.cell(1, col_idx, loc.get("level", ""))
            ws.cell(2, col_idx, loc.get("zone", ""))
            ws.cell(3, col_idx, loc.get("location_name", ""))

        # item rows
        row_num = 4

        for item_no in item_order:
            ws.cell(row_num, 1, item_no)
            ws.cell(row_num, 2, item_map[item_no]["item_name"])

            for col_idx, loc in enumerate(selected_locations, start=3):
                location_id = loc["location_id"]

                qty = qty_map.get(
                    (location_id, item_no),
                    0
                )

                ws.cell(
                    row_num,
                    col_idx,
                    qty if qty > 0 else ""
                )

            row_num += 1

        ws.freeze_panes = "C4"

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 35

        for col_idx in range(3, 3 + len(selected_locations)):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 8

    write_location_matrix_sheet(
        "Tower A",
        lambda loc: loc.get("tower") == "A"
    )

    headers = [
        "Item No",
        "Item Name",
        "New Stock",
        "Laundry Partner",
        "All Locations",
        "Condemn",
        "Grand Total"
    ]

    header_row = 6

    for col, header in enumerate(headers, start=1):
        summary_ws.cell(header_row, col, header)
        summary_ws.cell(header_row, col).font = Font(bold=True)

    NEW_STOCK_LOCS = {"LOC1600", "LOC1610"}

    LAUNDRY_PARTNER_LOCS = {"LOC1570"}

    CONDEMN_LOCS = {"LOC1560"}

    row_num = header_row + 1

    for item_no in item_order:

        item_name = item_map[item_no]["item_name"]

        new_stock = 0
        laundry_partner = 0
        condemn = 0
        all_locations = 0

        for (location_id, current_item_no), qty in qty_map.items():

            if current_item_no != item_no:
                continue

            if location_id in NEW_STOCK_LOCS:
                new_stock += qty

            elif location_id in LAUNDRY_PARTNER_LOCS:
                laundry_partner += qty

            elif location_id in CONDEMN_LOCS:
                condemn += qty

            else:
                all_locations += qty

        grand_total = (
            new_stock
            + laundry_partner
            + condemn
            + all_locations
        )
        
        summary_ws.cell(row_num, 1, item_no)
        summary_ws.cell(row_num, 2, item_name)

        summary_ws.cell(
            row_num, 3,
            new_stock if new_stock > 0 else ""
        )

        summary_ws.cell(
            row_num, 4,
            laundry_partner if laundry_partner > 0 else ""
        )

        summary_ws.cell(
            row_num, 5,
            all_locations if all_locations > 0 else ""
        )

        summary_ws.cell(
            row_num, 6,
            condemn if condemn > 0 else ""
        )

        summary_ws.cell(
            row_num, 7,
            grand_total if grand_total > 0 else ""
        )

        row_num += 1

    TOWER_B_EXCLUDE = {
        "LOC1550", "LOC1560", "LOC1570", "LOC1600", "LOC1620"
    }

    TOWER_C_EXCLUDE = {
        "LOC1610",
        "LOC1630", "LOC1640", "LOC1650", "LOC1660",
        "LOC1670", "LOC1680", "LOC1690", "LOC1700"
    }

    TROLLEY_LOCS = {
        "LOC1620",
        "LOC1630", "LOC1640", "LOC1650", "LOC1660",
        "LOC1670", "LOC1680", "LOC1690", "LOC1700"
    }

    write_location_matrix_sheet(
        "Tower B",
        lambda loc:
            loc.get("tower") == "B"
            and loc["location_id"] not in TOWER_B_EXCLUDE
    )

    write_location_matrix_sheet(
        "Tower C",
        lambda loc:
            loc.get("tower") == "C"
            and loc["location_id"] not in TOWER_C_EXCLUDE
    )

    write_location_matrix_sheet(
    "Tower C Item With Trolley",
        lambda loc:
            loc["location_id"] in TROLLEY_LOCS
    )

    write_location_matrix_sheet(
        "Linen Room",
        lambda loc:
            loc["location_id"] == "LOC1550"
    )

    write_location_matrix_sheet(
        "New Item",
        lambda loc:
            loc["location_id"] in {"LOC1600", "LOC1610"}
    )

    write_location_matrix_sheet(
        "Condemn",
        lambda loc:
            loc["location_id"] == "LOC1560"
    )

    write_location_matrix_sheet(
        "Laundry Partner",
        lambda loc:
            loc["location_id"] == "LOC1570"
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# ---------------------------
# Router
# ---------------------------
def router():
    page = st.session_state.get("page", "login")

    if page == "login":
        page_login()
    elif page == "home":
        page_home()
    elif page == "orders":
        page_orders()
    elif page == "order_detail":
        page_order_detail()
    elif page == "issue_stock":
        page_issue_stock()
    elif page == "stock_in":
        page_stock_in()
    elif page == "inventory":
        page_inventory()
    elif page == "monthly_report":
        page_monthly_report()
    elif page == "stock_card":
        page_stock_card()
    elif page == "system_tools":
        page_system_tools()
    elif page == "glo_gel_audits":
        page_glo_gel_audits()
    elif page == "glo_gel_audit_detail":
        page_glo_gel_audit_detail()
    elif page == "linen_inventory":
        page_linen_inventory()
    elif page == "linen_cycle_detail":
        page_linen_cycle_detail()
    elif page == "linen_count":
        page_linen_count()
    elif page == "linen_manage":
        page_linen_manage()
    else:
        st.session_state["page"] = "login"
        page_login()


# ---------------------------
# Main
# ---------------------------
def main():
    ensure_bootstrap()
    router()


if __name__ == "__main__":
    main()
